"""
Скрипт для создания раскладки изображения на 25 цветов
На основе wb_mozic
"""
import os
import sys
import math
import json
from PIL import Image, ImageDraw, ImageFont
import numpy as np
try:
    from skimage import color as skcolor
    HAS_SKIMAGE = True
except ImportError:
    HAS_SKIMAGE = False
    # Fallback: simple RGB to LAB conversion
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Импортируем конфигурацию
try:
    # Импортируем config из той же папки (slicer_utils)
    import importlib.util
    current_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(current_dir, 'config.py')
    if os.path.exists(config_path):
        spec = importlib.util.spec_from_file_location("slicer_config", config_path)
        config = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(config)
    else:
        import config
except (ImportError, AttributeError) as e:
    print(f"[WARNING] Файл config.py не найден или не содержит нужных атрибутов: {e}. Используются значения по умолчанию.")
    # Создаем объект с дефолтными значениями
    class config:
        NUM_COLORS = 25
        NUM_BLOCKS_WIDTH = 150
        NUM_BLOCKS_HEIGHT = 150
        BLOCKS_PER_PAGE_WIDTH = 56
        BLOCKS_PER_PAGE_HEIGHT = 85
        DPI = 300
        JPEG_QUALITY = 95
        MIN_CELL_SIZE_MM = 2
        SYMBOL_SIZE_RATIO = 0.7
        MIN_SYMBOL_SIZE_PT = 8
        PALETTE_SYMBOL_FONT_SIZE = 24
        PALETTE_NUMBER_FONT_SIZE = 30
        PALETTE_SYMBOLS_PER_ROW = 5
        ORGANIZER_FONT_SIZE = 36
        ORGANIZER_TITLE_FONT_SIZE = 42
        SHEET_DIR = "sheet"
        TASK_DIR = "task"
        STATIC_DIR = "static"
        SCALE_FONT_SIZE = 12
        INFO_FONT_SIZE = 18
        BAR_FONT_SIZE = 91
        TABLE_HEADER_FONT_SIZE = 12
        TABLE_DATA_FONT_SIZE = 15
        MIN_BLOCK_SIZE_PX = 10
        LAYOUT_SYMBOL_FONT_SCALE = 1.2
        MIN_LAYOUT_SYMBOL_FONT_SIZE = 20
        PALETTE_EXTRACTION_MAX_SIZE = None

try:
    from PyPDF2 import PdfMerger
except ImportError:
    try:
        from pypdf import PdfMerger
    except ImportError:
        PdfMerger = None
        print("[WARNING] PyPDF2 или pypdf не установлены. Объединение PDF будет недоступно.")

try:
    import pandas as pd
except ImportError:
    pd = None


def add_copyright_to_pdf(c, use_dmc=False, brand='marfa'):
    """
    Добавляет копирайт внизу слева на странице PDF маленьким шрифтом
    Args:
        c: объект canvas.Canvas
        use_dmc: если True, бренд "(c) Lilu&Stitch"
        brand: 'marfa' — Марфа Иголкина, 'mulen' — Мулен стиль (игнорируется при use_dmc=True)
    """
    if use_dmc:
        copyright_text = "\u00A9 Lilu&Stitch"
    elif brand == 'mulen':
        copyright_text = "\u00A9 Мулен стиль"
    else:
        copyright_text = "\u00A9 Марфа Иголкина"
    copyright_font_size = 9  # Размер шрифта копирайта
    copyright_margin = 5 * mm  # Отступ снизу
    copyright_left_margin = 5 * mm  # Отступ слева
    
    # Используем наш шрифт MontserratSemiBold (как на других страницах PDF)
    copyright_font_name, actual_font_size = get_scale_font(copyright_font_size)
    c.setFont(copyright_font_name, actual_font_size)
    c.setFillColorRGB(0, 0, 0)  # Черный цвет
    
    # Рисуем копирайт внизу слева
    c.drawString(copyright_left_margin, copyright_margin, copyright_text)


def set_pdf_metadata(canvas_obj, title=None):
    """
    Устанавливает метаданные PDF файла
    Args:
        canvas_obj: объект canvas.Canvas
        title: заголовок документа (опционально)
    """
    canvas_obj.setTitle(title or "Схема для вышивания")
    canvas_obj.setAuthor("Zhdanov V.Y, Zhdanov I.Y.")
    canvas_obj.setSubject("Схема для вышивания крестом")
    canvas_obj.setCreator("NexelSoftware")
    canvas_obj.setKeywords("вышивание, схема, крестом, embroidery, cross stitch")
    # Авторское право устанавливается через Producer
    canvas_obj.setProducer("Zhdanov V.Y, Zhdanov I.Y. - Copyright")


def set_pdf_merger_metadata(merger, title=None):
    """
    Устанавливает метаданные для объединенного PDF файла
    Args:
        merger: объект PdfMerger
        title: заголовок документа (опционально)
    """
    metadata = {
        '/Title': title or "Схема для вышивания",
        '/Author': 'Zhdanov V.Y, Zhdanov I.Y.',
        '/Subject': 'Схема для вышивания крестом',
        '/Creator': 'NexelSoftware',
        '/Keywords': 'вышивание, схема, крестом, embroidery, cross stitch',
        '/Producer': 'Zhdanov V.Y, Zhdanov I.Y. - Copyright'
    }
    try:
        # PyPDF2 использует add_metadata
        if hasattr(merger, 'add_metadata'):
            merger.add_metadata(metadata)
        # pypdf использует metadata атрибут
        elif hasattr(merger, 'metadata'):
            merger.metadata = metadata
    except Exception as e:
        print(f"[WARNING] Не удалось установить метаданные для объединенного PDF: {e}")


# Флаг для отслеживания, было ли выведено предупреждение о piexif
_piexif_warning_shown = False

def set_jpg_metadata(image_obj, title=None):
    """
    Устанавливает метаданные JPG изображения (EXIF/IPTC)
    Args:
        image_obj: объект PIL.Image
        title: заголовок документа (опционально)
    Returns:
        dict: Словарь с метаданными для сохранения
    """
    global _piexif_warning_shown
    try:
        # Пробуем использовать piexif для более полной поддержки метаданных
        try:
            import piexif
            
            # Создаем EXIF данные
            exif_dict = {}
            
            # 0th IFD (Image)
            exif_dict['0th'] = {}
            exif_dict['0th'][piexif.ImageIFD.ImageDescription] = (title or "Схема для вышивания").encode('utf-8')
            exif_dict['0th'][piexif.ImageIFD.Artist] = "Zhdanov V.Y, Zhdanov I.Y.".encode('utf-8')
            exif_dict['0th'][piexif.ImageIFD.Copyright] = "Zhdanov V.Y, Zhdanov I.Y. - Copyright".encode('utf-8')
            exif_dict['0th'][piexif.ImageIFD.Software] = "NexelSoftware - Редактор схемы вышивки".encode('utf-8')
            
            # EXIF IFD
            exif_dict['Exif'] = {}
            
            # Преобразуем в байты
            exif_bytes = piexif.dump(exif_dict)
            return {'exif': exif_bytes}
        except ImportError:
            # Если piexif не установлен, выводим предупреждение только один раз
            if not _piexif_warning_shown:
                print("[WARNING] Библиотека piexif не установлена. Метаданные JPG будут сохранены без EXIF данных.")
                print("[INFO] Для полной поддержки метаданных установите: pip install piexif")
                _piexif_warning_shown = True
            # Возвращаем пустой dict - метаданные не будут добавлены
            return {}
    except Exception as e:
        if not _piexif_warning_shown:
            print(f"[WARNING] Не удалось создать метаданные для JPG: {e}")
            _piexif_warning_shown = True
        return {}


def save_jpg_with_metadata(image_obj, output_path, quality=95, dpi=None, title=None):
    """
    Сохраняет JPG изображение с метаданными
    Args:
        image_obj: объект PIL.Image
        output_path: путь для сохранения
        quality: качество JPEG (1-100)
        dpi: разрешение (кортеж (x, y) или число)
        title: заголовок документа (опционально)
    """
    # Получаем метаданные
    metadata = set_jpg_metadata(image_obj, title)
    
    # Подготавливаем параметры для сохранения
    save_kwargs = {
        'format': 'JPEG',
        'quality': quality,
        'optimize': True
    }
    
    # Добавляем DPI если указан
    if dpi:
        if isinstance(dpi, (int, float)):
            save_kwargs['dpi'] = (int(dpi), int(dpi))
        else:
            save_kwargs['dpi'] = dpi
    
    # Добавляем метаданные если они есть
    if metadata:
        save_kwargs.update(metadata)
    
    # Сохраняем изображение
    image_obj.save(output_path, **save_kwargs)

# Специальные символы для обозначения цветов
# Только символы, которые корректно отображаются в PDF (без черных квадратов)
# Геометрические символы и базовые символы, поддерживаемые стандартными PDF шрифтами
SPECIAL_CHARS = [
    # Unicode символы для красивого отображения в PDF (используются первыми)
    # Треугольники и стрелки
    
    # Звездочки и фигуры
    '§', '*', '↑', '†', '>', '⬛', '√', '✪', '+', '♥', '÷', '➡', '⊂', '//', '⊕', '∴', '|', '♠', '◊', '✺', '∫', 'Ξ', 'ψ', 'Γ', '✿', '❀', '❁', '❂',
    # Геометрические фигуры
    '●', '○', '◆', '■', '□', '▲', '▼',  '◉', '◈', '◊', '◌', '◍', '◐', '◑', '◒', '◓', '◔', '◕', '◖', '◗', '◘', '◙', '◚', '◛', '◜', '◝', '◞', '◟', '◠', '◡', '◢', '◣', '◤', '◥',
    # Символы из KoiFishesWaterLilyPatternKeeper.pdf
    'm', '>', '<',
    '甲', 'ㄱ', 'ㄷ', 'ㅁ', 'ㅍ',  # Корейские символы
    '//', '□',  # Специальные символы
    # Дополнительные геометрические символы (хорошо поддерживаются)
    '', '◐', '◑', '◒', '◓', '◔', '◕', '◖', '◗', '◘', '◙', '◚',
    '◛', '◜', '◝', '◞', '◟', '◠', '◡', '◢', '◣',
    '◥', '◦', '◧', '◨', '◩', '◪', '◫', '◬', '◭', '◮',
    '◯', '◰', '◱', '◲', '◳', '◴', '◵', '◶', '◷',
    # Дополнительные символы из wb_mozic (базовые, без эмодзи)
    '♦', '♥', '♣', '♨',
    # Стрелки и направляющие
    '⬆', '⬇', '⬅', '➡', '⬛',
    # Простые символы
    '©', '¨',
    # Дополнительные базовые ASCII символы (гарантированно работают в PDF)
    'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'Y', 'Z',
    'a', 'b', 'f', 'h', 'l', 'n', 't', 'u', 'v', 'w', 'x', 'y', 'z',
    # Цифры и знаки
    '+', '-', '=', '/', '\\', '_', '.', ',', ':', ';', '!', '?', '%', '&', '$', '#',
    # Математические символы (хорошо поддерживаются в PDF)
    '×', '÷', '±', '≠', '≤', '≥', '≈', '∞', '∑', '∏', '√', '∫', '∆', '∇',
    # Дополнительные геометрические символы
    '▪', '▫', '▬', '▭', '▮', '▯', '▰', '▱', '▲', '△', '▴', '▵', '▶', '▷', '▸', '▹',
    '▼', '▽', '▾', '▿', '◀', '◁', '◂', '◃', '◆', '◇', '◈', '◉', '◊', '○', '●', '◐', '◑', '◒', '◓',
    '◔', '◕', '◖', '◗', '◘', '◙', '◚', '◛', '◜', '◝', '◞', '◟', '', '◡', '◢', '◣', '◤', '◥',
    '◦', '◧', '◨', '◩', '◪', '◫', '◬', '◭', '◮', '◯', '◰', '◱', '◲', '◳', '◴', '◵', '◶', '◷',
    # Блоковые символы
    '▁', '▂', '▃', '▄', '▅', '▆', '▇', '█', '▉', '▊', '▋', '▌', '▍', '▎', '▏',
    '▐', '░', '▒', '▓', '▔', '▕', '▖', '▗', '▘', '▙', '▚', '▛', '▜', '▝', '▞', '▟',
    # Дополнительные стрелки
    '←', '→', '↑', '↓', '↔', '↕', '↖', '↗', '↘', '↙', '⇐', '⇒', '⇑', '⇓', '⇔', '⇕',
    # Дополнительные звездочки и символы
    '★', '☆', '✪', '✫', '✬', '✭', '✮', '✯', '✰', '✱', '✲', '✳', '✴', '✵', '✶', '✷', '✸', '✹', '✺', '✻', '✼', '✽', '✾', '✿', '❀', '❁', '❂', '❃', '❄', '❅', '❆', '❇', '❈', '❉', '❊', '❋',
    # Дополнительные карточные масти
    '♤', '♧', '♡', '♢',
    # Дополнительные символы
    '§', '¶', '†', '‡', '•', '…', '‰', '′', '″', '‴', '※', '⁂', '⁎', '⁑', '⁒', '⁓', '⁕', '⁖', '⁗', '⁘', '⁙', '⁚', '⁛', '⁜', '⁝', '⁞',
    # Греческие буквы (базовые, хорошо поддерживаются)
    'α', 'β', 'γ', 'δ', 'ε', 'ζ', 'η', 'θ', 'ι', 'κ', 'λ', 'μ', 'ν', 'ξ', 'ο', 'π', 'ρ', 'σ', 'τ', 'υ', 'φ', 'χ', 'ψ', 'ω',
    'Α', 'Β', 'Γ', 'Δ', 'Ε', 'Ζ', 'Η', 'Θ', 'Ι', 'Κ', 'Λ', 'Μ', 'Ν', 'Ξ', 'Ο', 'Π', 'Ρ', 'Σ', 'Τ', 'Υ', 'Φ', 'Χ', 'Ψ', 'Ω',
    # Дополнительные специальные символы
    '°', '²', '³', '¹', '½', '¾', '⅓', '⅔', '⅛', '⅜', '⅝', '⅞',
    # Символы валют и другие
    '€', '£', '¥', '¢', '¤', '₽',
    # Дополнительные пунктуационные символы
    '«', '»', '„', '‚', '‹', '›', '‛', '‚', '„', '…', '‧', '‥', '‧', '‥'
]


def extract_dominant_colors(image_path, num_colors=None):
    """
    Извлекает доминирующие цвета из изображения используя квантование цветов.
    Использует оригинальный размер изображения или ограничивает его максимальным размером
    из конфига (PALETTE_EXTRACTION_MAX_SIZE) для больших изображений.
    """
    if num_colors is None:
        num_colors = config.NUM_COLORS
    
    # Загружаем изображение
    img = Image.open(image_path)
    img = img.convert("RGB")
    
    # Используем оригинальный размер или ограничиваем максимальным размером из конфига
    original_width, original_height = img.size
    max_size = getattr(config, 'PALETTE_EXTRACTION_MAX_SIZE', None)
    
    if max_size is not None and max_size > 0:
        # Ограничиваем максимальным размером по большей стороне
        max_dimension = max(original_width, original_height)
        if max_dimension > max_size:
            # Вычисляем коэффициент масштабирования
            scale = max_size / max_dimension
            new_width = int(original_width * scale)
            new_height = int(original_height * scale)
            img_small = img.resize((new_width, new_height), Image.Resampling.NEAREST)
            print(f"[PALETTE] Изображение уменьшено для извлечения палитры: {original_width}x{original_height} -> {new_width}x{new_height} (макс. размер: {max_size})")
        else:
            # Используем оригинальный размер
            img_small = img
            print(f"[PALETTE] Используется оригинальный размер изображения для извлечения палитры: {original_width}x{original_height}")
    else:
        # Используем оригинальный размер (без ограничений)
        img_small = img
        print(f"[PALETTE] Используется оригинальный размер изображения для извлечения палитры: {original_width}x{original_height}")
    
    # Используем метод quantize для уменьшения количества цветов
    # Сначала увеличиваем количество цветов, чтобы получить более точную палитру
    # Используем MEDIANCUT для лучшего распределения цветов (принцип из Pixel Logic)
    quantized = img_small.quantize(colors=num_colors * 3, method=Image.Quantize.MEDIANCUT)
    
    # Получаем палитру
    palette_colors = quantized.getpalette()
    
    # Извлекаем уникальные цвета из палитры
    colors = []
    for i in range(0, len(palette_colors), 3):
        if i + 2 < len(palette_colors):
            r, g, b = palette_colors[i], palette_colors[i + 1], palette_colors[i + 2]
            if (r, g, b) not in colors:
                colors.append((r, g, b))
    
    # Если получилось больше цветов, чем нужно, выбираем наиболее частые
    if len(colors) > num_colors:
        # Подсчитываем частоту каждого цвета
        pixel_array = np.array(quantized)
        color_counts = {}
        for color_idx in np.unique(pixel_array):
            if color_idx < len(colors):
                color = colors[color_idx]
                count = np.sum(pixel_array == color_idx)
                color_counts[color] = count
        
        # Сортируем по частоте и берем топ-N
        sorted_colors = sorted(color_counts.items(), key=lambda x: x[1], reverse=True)
        colors = [color for color, count in sorted_colors[:num_colors]]
    
    # Если получилось меньше цветов, дополняем равномерно распределенными
    if len(colors) < num_colors:
        # Добавляем цвета из равномерного распределения в RGB пространстве
        step = int(256 / ((num_colors - len(colors)) ** (1/3)))
        for r in range(0, 256, step):
            for g in range(0, 256, step):
                for b in range(0, 256, step):
                    if len(colors) >= num_colors:
                        break
                    if (r, g, b) not in colors:
                        colors.append((r, g, b))
                if len(colors) >= num_colors:
                    break
            if len(colors) >= num_colors:
                break
    
    palette = colors[:num_colors]
    
    # Оптимизируем палитру (удаляем похожие цвета, упорядочиваем)
    optimize_palette_flag = getattr(config, 'OPTIMIZE_PALETTE', True)
    if optimize_palette_flag:
        palette = optimize_palette(palette, min_contrast=getattr(config, 'PALETTE_MIN_CONTRAST', 30))
        # Если после оптимизации цветов стало меньше, дополняем
        if len(palette) < num_colors:
            print(f"[PALETTE] После оптимизации осталось {len(palette)} цветов, дополняем до {num_colors}")
            # Дополняем равномерно распределенными цветами
            step = int(256 / ((num_colors - len(palette)) ** (1/3)))
            for r in range(0, 256, step):
                for g in range(0, 256, step):
                    for b in range(0, 256, step):
                        if len(palette) >= num_colors:
                            break
                        new_color = (r, g, b)
                        # Проверяем, что новый цвет достаточно отличается от существующих
                        if all(color_distance(new_color, c) >= getattr(config, 'PALETTE_MIN_CONTRAST', 30) for c in palette):
                            palette.append(new_color)
                    if len(palette) >= num_colors:
                        break
                if len(palette) >= num_colors:
                    break
        palette = palette[:num_colors]
    
    return palette


def optimize_palette(palette, min_contrast=30):
    """
    Оптимизирует палитру, удаляя слишком похожие цвета и упорядочивая по яркости.
    Принцип из Pixel Logic: гармоничная палитра с достаточным контрастом.
    """
    if not palette:
        return palette
    
    # Удаляем слишком похожие цвета
    optimized = []
    for color in palette:
        is_unique = True
        for existing in optimized:
            if color_distance(color, existing) < min_contrast:
                is_unique = False
                break
        if is_unique:
            optimized.append(color)
    
    # Упорядочиваем по яркости (L в LAB или простая формула)
    def get_luminance(rgb):
        # Формула яркости для восприятия человеком
        return 0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]
    
    optimized.sort(key=get_luminance)
    return optimized


def cleanup_artifacts(block_colors, num_blocks, min_cluster_size=2):
    """
    Удаляет одиночные блоки и мелкие артефакты (принцип Clean Up из Pixel Logic).
    Улучшает качество пикселизации, удаляя шум и неровности.
    """
    from collections import Counter
    
    cleaned = block_colors.copy()
    changed_count = 0
    
    # Находим одиночные блоки (блоки, у которых все соседи другого цвета)
    for y in range(num_blocks):
        for x in range(num_blocks):
            current = tuple(block_colors[y, x])
            neighbors = []
            for dy in [-1, 0, 1]:
                for dx in [-1, 0, 1]:
                    if dy == 0 and dx == 0:
                        continue
                    ny, nx = y + dy, x + dx
                    if 0 <= ny < num_blocks and 0 <= nx < num_blocks:
                        neighbors.append(tuple(block_colors[ny, nx]))
            
            # Если все соседи другого цвета - это артефакт
            if neighbors and all(n != current for n in neighbors):
                # Заменяем на наиболее частый цвет соседей
                neighbor_counts = Counter(neighbors)
                if neighbor_counts:
                    most_common = neighbor_counts.most_common(1)[0][0]
                    cleaned[y, x] = most_common
                    changed_count += 1
    
    if changed_count > 0:
        print(f"[CLEANUP] Удалено {changed_count} артефактов (одиночных блоков)")
    
    return cleaned


def get_optimal_text_color(background_color):
    """
    Выбирает оптимальный цвет текста (черный или белый) для максимальной читаемости.
    Принцип из Pixel Logic: обеспечение достаточного контраста для читаемости.
    """
    # Вычисляем яркость фона по формуле восприятия человеком
    luminance = 0.299 * background_color[0] + 0.587 * background_color[1] + 0.114 * background_color[2]
    
    # Если фон светлый - черный текст, если темный - белый
    return (0, 0, 0) if luminance > 128 else (255, 255, 255)


def rgb_to_lab(rgb):
    """
    Конвертирует RGB в LAB цветовое пространство для перцептуально точного сравнения.
    Использует упрощенную формулу, если skimage недоступен.
    """
    r, g, b = rgb[0] / 255.0, rgb[1] / 255.0, rgb[2] / 255.0
    
    # Гамма-коррекция
    def gamma_correct(x):
        if x > 0.04045:
            return ((x + 0.055) / 1.055) ** 2.4
        return x / 12.92
    
    r = gamma_correct(r)
    g = gamma_correct(g)
    b = gamma_correct(b)
    
    # Конвертация в XYZ
    x = (r * 0.4124 + g * 0.3576 + b * 0.1805) / 0.95047
    y = (r * 0.2126 + g * 0.7152 + b * 0.0722) / 1.00000
    z = (r * 0.0193 + g * 0.1192 + b * 0.9505) / 1.08883
    
    # Конвертация в LAB
    def f(t):
        if t > 0.008856:
            return t ** (1.0/3.0)
        return (7.787 * t) + (16.0/116.0)
    
    fx, fy, fz = f(x), f(y), f(z)
    
    L = (116.0 * fy) - 16.0
    lab_a = 500.0 * (fx - fy)
    lab_b = 200.0 * (fy - fz)
    
    return (L, lab_a, lab_b)


def color_distance(c1, c2):
    """
    Вычисляет перцептуально точное расстояние между двумя цветами в LAB пространстве.
    Это дает более точные результаты, чем простое RGB расстояние, так как учитывает
    особенности человеческого восприятия цветов (принцип из Pixel Logic).
    """
    if HAS_SKIMAGE:
        # Используем skimage для точной конвертации
        try:
            lab1 = skcolor.rgb2lab(np.array([[c1]]) / 255.0)[0, 0]
            lab2 = skcolor.rgb2lab(np.array([[c2]]) / 255.0)[0, 0]
            # Delta E (CIE76) - перцептуальное расстояние
            return np.sqrt(sum((a - b) ** 2 for a, b in zip(lab1, lab2)))
        except:
            pass
    
    # Fallback: используем упрощенную конвертацию
    lab1 = rgb_to_lab(c1)
    lab2 = rgb_to_lab(c2)
    # Delta E (CIE76) - перцептуальное расстояние в LAB
    return np.sqrt(sum((a - b) ** 2 for a, b in zip(lab1, lab2)))


def apply_block_level_dithering(block_colors, num_blocks, palette):
    """
    Оптимизированный дизеринг на уровне блоков (быстрее, чем пиксельный).
    Применяет упрощенный алгоритм дизеринга для создания плавных градиентов
    с ограниченной палитрой (принцип дизеринга из Pixel Logic).
    Это создает иллюзию дополнительных оттенков через паттерны блоков.
    """
    result = block_colors.copy().astype(np.float32)
    
    for y in range(num_blocks):
        for x in range(num_blocks):
            old_color = result[y, x].copy()
            
            # Находим ближайший цвет из палитры
            closest_color = min(palette, key=lambda c: color_distance(old_color.astype(int), c))
            result[y, x] = np.array(closest_color, dtype=np.float32)
            
            # Вычисляем ошибку квантования
            error = old_color - result[y, x]
            
            # Распределяем ошибку на соседние блоки (упрощенный Floyd-Steinberg)
            # Только на соседние блоки справа и снизу для скорости
            if x + 1 < num_blocks:
                result[y, x + 1] += error * 0.5  # 50% ошибки вправо
            if y + 1 < num_blocks:
                result[y + 1, x] += error * 0.3  # 30% ошибки вниз
                if x + 1 < num_blocks:
                    result[y + 1, x + 1] += error * 0.2  # 20% ошибки по диагонали
    
    # Ограничиваем значения в диапазоне [0, 255]
    result = np.clip(result, 0, 255).astype(np.uint8)
    return result


def fix_colors_by_neighbors(block_colors, num_blocks, preserve_edges=True):
    """
    Улучшенное исправление цветов блоков с сохранением краев (принцип из Pixel Logic).
    Исправляет цвета блоков по правилам:
    - Если по диагоналям и по прямой у кубика нет похожего цвета, то его цвет меняется на тот, который по прямой.
    - Сохраняет резкие края (edges) для лучшей читаемости.
    Улучшенная версия с более точной детекцией краев.
    """
    from collections import Counter
    
    fixed_colors = block_colors.copy()
    changed_count = 0
    
    # Улучшенная детекция краев с использованием всех 8 соседей
    edge_map = np.zeros((num_blocks, num_blocks), dtype=bool)
    if preserve_edges:
        # Получаем порог из конфига или используем более чувствительный по умолчанию
        edge_threshold = getattr(config, 'EDGE_DETECTION_THRESHOLD', 20)  # Более чувствительный порог
        
        # Первый проход: детекция краев с учетом всех 8 соседей
        for y in range(num_blocks):
            for x in range(num_blocks):
                current = tuple(fixed_colors[y, x])
                
                # Проверяем все 8 соседей (включая диагонали)
                neighbors = []
                for dy in [-1, 0, 1]:
                    for dx in [-1, 0, 1]:
                        if dy == 0 and dx == 0:
                            continue  # Пропускаем сам блок
                        ny, nx = y + dy, x + dx
                        if 0 <= ny < num_blocks and 0 <= nx < num_blocks:
                            neighbors.append(tuple(fixed_colors[ny, nx]))
                
                if neighbors:
                    # Вычисляем максимальную разницу с соседями
                    max_diff = max(color_distance(current, n) for n in neighbors)
                    
                    # Также проверяем среднюю разницу для более надежной детекции
                    avg_diff = sum(color_distance(current, n) for n in neighbors) / len(neighbors)
                    
                    # Блок считается краем, если:
                    # 1. Максимальная разница превышает порог ИЛИ
                    # 2. Средняя разница превышает порог (для более чувствительной детекции)
                    edge_map[y, x] = (max_diff > edge_threshold) or (avg_diff > edge_threshold * 0.7)
        
        # Второй проход: морфологическое закрытие для заполнения пропусков
        # Расширяем края на соседние блоки, если они тоже имеют значительную разницу
        edge_map_expanded = edge_map.copy()
        for y in range(num_blocks):
            for x in range(num_blocks):
                if not edge_map[y, x]:  # Если это не край
                    current = tuple(fixed_colors[y, x])
                    # Проверяем соседей, которые уже определены как края
                    edge_neighbors = []
                    for dy in [-1, 0, 1]:
                        for dx in [-1, 0, 1]:
                            if dy == 0 and dx == 0:
                                continue
                            ny, nx = y + dy, x + dx
                            if 0 <= ny < num_blocks and 0 <= nx < num_blocks:
                                if edge_map[ny, nx]:  # Если сосед - край
                                    edge_neighbors.append(tuple(fixed_colors[ny, nx]))
                    
                    # Если рядом есть края и разница с ними значительна, тоже помечаем как край
                    if edge_neighbors:
                        min_diff_to_edge = min(color_distance(current, n) for n in edge_neighbors)
                        if min_diff_to_edge > edge_threshold * 0.5:  # Более мягкий порог для расширения
                            edge_map_expanded[y, x] = True
        
        edge_map = edge_map_expanded
    
    for y in range(num_blocks):
        for x in range(num_blocks):
            # Пропускаем края, если включено сохранение краев
            if preserve_edges and edge_map[y, x]:
                continue
                
            current_color = tuple(fixed_colors[y, x])
            
            # Собираем цвета по диагоналям (4 диагональных соседа)
            diagonal_colors = []
            for dy, dx in [(-1, -1), (-1, 1), (1, -1), (1, 1)]:
                ny, nx = y + dy, x + dx
                if 0 <= ny < num_blocks and 0 <= nx < num_blocks:
                    diagonal_colors.append(tuple(fixed_colors[ny, nx]))
            
            # Собираем цвета по прямым линиям (вертикаль и горизонталь)
            straight_colors = []
            for dy, dx in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                ny, nx = y + dy, x + dx
                if 0 <= ny < num_blocks and 0 <= nx < num_blocks:
                    straight_colors.append(tuple(fixed_colors[ny, nx]))
            
            # Проверяем, есть ли похожий цвет по диагоналям (точно такой же)
            has_similar_diagonal = any(dc == current_color for dc in diagonal_colors)
            
            # Проверяем, есть ли похожий цвет по прямым (точно такой же)
            has_similar_straight = any(sc == current_color for sc in straight_colors)
            
            # Если ни по диагоналям, ни по прямой нет похожего цвета
            if not has_similar_diagonal and not has_similar_straight:
                # Заменяем на цвет по прямой (берем наиболее частый цвет среди прямых соседей)
                if straight_colors:
                    straight_counts = Counter(straight_colors)
                    if straight_counts:
                        most_common_straight = straight_counts.most_common(1)[0][0]
                        fixed_colors[y, x] = most_common_straight
                        changed_count += 1
    
    return fixed_colors


# Глобальные переменные для отслеживания регистрации шрифтов
_scale_font_registered = False
_bar_font_registered = False
_info_font_registered = False

def get_scale_font(font_size):
    """
    Загружает шрифт из static/fonts для цифр в шкале
    Возвращает имя зарегистрированного шрифта и размер
    """
    global _scale_font_registered
    font_name = "MontserratSemiBold"
    try:
        from utils.path_utils import get_static_path
        font_path = get_static_path("fonts/Montserrat-SemiBold.ttf")
    except ImportError:
        font_path = os.path.join("static", "fonts", "Montserrat-SemiBold.ttf")
    
    # Пробуем загрузить шрифт из static/fonts (регистрируем только один раз)
    if not _scale_font_registered and os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont(font_name, font_path))
            _scale_font_registered = True
        except Exception as e:
            print(f"[WARNING] Не удалось загрузить шрифт из {font_path}: {e}")
            return "Helvetica", font_size
    
    # Если шрифт зарегистрирован, используем его, иначе Helvetica
    if _scale_font_registered:
        return font_name, font_size
    else:
        return "Helvetica", font_size


def get_bar_font(font_size):
    """
    Загружает шрифт из static/fonts для текста на баре (жирный)
    Возвращает имя зарегистрированного шрифта и размер
    """
    global _bar_font_registered
    font_name = "MontserratSemiBold"
    try:
        from utils.path_utils import get_static_path
        font_path = get_static_path("fonts/Montserrat-SemiBold.ttf")
    except ImportError:
        font_path = os.path.join("static", "fonts", "Montserrat-SemiBold.ttf")
    
    if not _bar_font_registered and os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont(font_name, font_path))
            _bar_font_registered = True
        except Exception as e:
            print(f"[WARNING] Не удалось загрузить шрифт из {font_path}: {e}")
            return "Helvetica-Bold", font_size
    
    if _bar_font_registered or _scale_font_registered:
        return font_name, font_size
    else:
        return "Helvetica-Bold", font_size


def get_info_font(font_size):
    """
    Загружает шрифт из static/fonts для информационного текста
    Возвращает имя зарегистрированного шрифта и размер
    """
    global _info_font_registered
    # Пробуем разные шрифты в порядке приоритета
    try:
        from utils.path_utils import get_static_path
        font_options = [
            ("MontserratSemiBold", get_static_path("fonts/Montserrat-SemiBold.ttf")),
            ("MontserratRegular", get_static_path("fonts/MontserratAlternates-Regular.otf")),
            ("MontserratSemiBoldAlt", get_static_path("fonts/MontserratAlternates-SemiBold.otf")),
        ]
    except ImportError:
        font_options = [
            ("MontserratSemiBold", os.path.join("static", "fonts", "Montserrat-SemiBold.ttf")),
            ("MontserratRegular", os.path.join("static", "fonts", "MontserratAlternates-Regular.otf")),
            ("MontserratSemiBoldAlt", os.path.join("static", "fonts", "MontserratAlternates-SemiBold.otf")),
        ]
    
    font_name = None
    font_path = None
    
    for name, path in font_options:
        if os.path.exists(path):
            font_name = name
            font_path = path
            break
    
    if font_name and font_path:
        if not _info_font_registered:
            try:
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                _info_font_registered = True
            except Exception as e:
                print(f"[WARNING] Не удалось загрузить шрифт из {font_path}: {e}")
                return "Helvetica", font_size
        
        if _info_font_registered:
            return font_name, font_size
    
    # Fallback к стандартному шрифту
    print(f"[WARNING] Шрифт для информационного текста не найден, используем Helvetica")
    return "Helvetica", font_size


def get_font(size):
    """Универсальная функция для загрузки шрифтов на разных ОС"""
    import platform
    
    # Для Windows используем шрифты, которые поддерживают Unicode символы
    if platform.system() == "Windows":
        font_candidates = [
            ("C:/Windows/Fonts/segoeui.ttf", "Windows Segoe UI"),
            ("C:/Windows/Fonts/seguiemj.ttf", "Windows Segoe UI Emoji"),
            ("C:/Windows/Fonts/arial.ttf", "Windows Arial"),
            ("C:/Windows/Fonts/calibri.ttf", "Windows Calibri"),
            ("seguisym.ttf", "Windows Symbol"),
            ("arial.ttf", "Windows Arial"),
        ]
    else:
        font_candidates = [
            ("DejaVuSans.ttf", "Ubuntu/Linux DejaVu"),
            ("liberation-sans.ttf", "Ubuntu Liberation"),
            ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "Ubuntu DejaVu (полный путь)"),
            ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", "Ubuntu Liberation (полный путь)"),
            ("/System/Library/Fonts/Arial.ttf", "macOS Arial"),
        ]
    
    for font_path, description in font_candidates:
        try:
            font = ImageFont.truetype(font_path, size)
            return font
        except (OSError, IOError) as e:
            continue
    
    # Если ни один шрифт не загрузился, используем дефолтный
    return ImageFont.load_default()


def create_color_layout(input_image_path, output_image_path, num_colors=None, num_blocks_width=None, num_blocks_height=None):
    """
    Создает раскладку изображения с ограниченной палитрой цветов
    """
    if num_colors is None:
        num_colors = config.NUM_COLORS
    if num_blocks_width is None:
        num_blocks_width = config.NUM_BLOCKS_WIDTH
    if num_blocks_height is None:
        num_blocks_height = config.NUM_BLOCKS_HEIGHT
    """
    Создает раскладку изображения с использованием ограниченной палитры цветов
    """
    # Извлекаем доминирующие цвета
    palette = extract_dominant_colors(input_image_path, num_colors)
    
    # Загружаем исходное изображение
    img = Image.open(input_image_path)
    img = img.convert("RGB")
    
    # Делаем изображение квадратным для создания квадратных блоков
    width, height = img.size
    min_side = min(width, height)
    
    # Обрезаем изображение до квадрата (центрируем)
    if width > height:
        # Обрезаем по ширине
        left = (width - height) // 2
        img_cropped = img.crop((left, 0, left + height, height))
    elif height > width:
        # Обрезаем по высоте
        top = (height - width) // 2
        img_cropped = img.crop((0, top, width, top + width))
    else:
        # Уже квадратное
        img_cropped = img
    
    # Вычисляем размер квадратного блока на основе квадратного изображения
    # Увеличиваем размер, чтобы символы были видны (минимум 10 пикселей на блок)
    square_size = min_side
    num_blocks = max(num_blocks_width, num_blocks_height)
    block_size = max(square_size // num_blocks, 10)  # Минимум 10 пикселей
    
    # Пересчитываем размер раскладки с учетом минимального размера блока
    layout_size = num_blocks * block_size
    
    # Изменяем размер квадратного изображения под размеры блоков
    img_resized = img_cropped.resize((layout_size, layout_size), Image.Resampling.NEAREST)
    
    # Создаем соответствие между цветами и символами
    char_list = SPECIAL_CHARS * ((len(palette) // len(SPECIAL_CHARS)) + 1)
    color_to_char = {color: char_list[i] for i, color in enumerate(palette)}
    
    # Получаем шрифт для символов (увеличиваем минимальный размер)
    font_size = max(int(block_size * config.LAYOUT_SYMBOL_FONT_SCALE), config.MIN_LAYOUT_SYMBOL_FONT_SIZE)
    font = get_font(font_size)
    
    # Тестируем отрисовку символа (без вывода символа в консоль из-за проблем с кодировкой)
    test_symbol = SPECIAL_CHARS[0]
    test_img = Image.new("RGB", (50, 50), "white")
    test_draw = ImageDraw.Draw(test_img)
    try:
        test_draw.text((10, 10), test_symbol, fill="green", font=font)
    except Exception as e:
        # Пробуем дефолтный шрифт
        default_font = ImageFont.load_default()
        test_draw.text((10, 10), test_symbol, fill="green", font=default_font)
    
    # Создаем схему раскладки (квадратное изображение)
    layout = Image.new("RGB", (layout_size, layout_size), "white")
    draw = ImageDraw.Draw(layout)
    
    # Сначала собираем все цвета блоков в массив
    block_colors = np.zeros((num_blocks, num_blocks, 3), dtype=np.uint8)
    total_blocks = num_blocks * num_blocks
    processed = 0
    
    for x in range(num_blocks):
        for y in range(num_blocks):
            left = x * block_size
            upper = y * block_size
            right = left + block_size
            lower = upper + block_size
            
            # Извлекаем блок
            block = img_resized.crop((left, upper, right, lower))
            
            # Вычисляем средний цвет блока (используем медиану для лучшей устойчивости к выбросам)
            block_array = np.array(block)
            # Используем медиану вместо среднего для лучшей устойчивости
            median_color = tuple(np.median(block_array.reshape(-1, 3), axis=0).astype(int))
            
            # Находим ближайший цвет из палитры (используя перцептуальное расстояние)
            closest_color = min(palette, key=lambda c: color_distance(median_color, c))
            
            # Сохраняем цвет блока
            block_colors[y, x] = closest_color
            
            processed += 1
    
    # Применяем дизеринг на уровне блоков (оптимизированная версия - намного быстрее!)
    # Это создает более плавные переходы цветов между блоками
    enable_dithering = getattr(config, 'ENABLE_DITHERING', True)
    if enable_dithering:
        block_colors = apply_block_level_dithering(block_colors, num_blocks, palette)
    
    # Применяем исправление цветов на основе соседей с сохранением краев
    preserve_edges = getattr(config, 'PRESERVE_EDGES', True)
    block_colors = fix_colors_by_neighbors(block_colors, num_blocks, preserve_edges=preserve_edges)
    
    # Применяем очистку артефактов (принцип Clean Up из Pixel Logic)
    cleanup_artifacts_flag = getattr(config, 'CLEANUP_ARTIFACTS', True)
    if cleanup_artifacts_flag:
        block_colors = cleanup_artifacts(block_colors, num_blocks)
    
    # Теперь рисуем блоки с исправленными цветами
    for x in range(num_blocks):
        for y in range(num_blocks):
            left = x * block_size
            upper = y * block_size
            right = left + block_size
            lower = upper + block_size
            
            # Получаем цвет из исправленного массива
            color = tuple(block_colors[y, x])
            
            # Рисуем блок с цветом из палитры (без символов)
            draw.rectangle([left, upper, right, lower], fill=color, outline="black", width=1)
    
    # Сохраняем результат без символов в формате JPG с метаданными
    layout_rgb = layout.convert("RGB")
    save_jpg_with_metadata(layout_rgb, output_image_path, quality=config.JPEG_QUALITY, title="Схема для вышивания")
    
    # НЕ создаем палитру здесь - она будет создана позже в main() с правильным color_to_char
    # create_palette_image(palette, color_to_char, output_image_path.replace('.', '_palette.'))
    
    return palette


def create_symbol_image(symbol, font_size, text_color="black", bg_color=None):
    """
    Создает изображение символа с прозрачным фоном
    Возвращает RGBA изображение с символом
    """
    import platform
    
    # Размер изображения для символа (достаточно большой для качества)
    img_size = font_size * 3
    symbol_img = Image.new("RGBA", (img_size, img_size), (0, 0, 0, 0))  # Прозрачный фон
    symbol_draw = ImageDraw.Draw(symbol_img)
    
    # Пробуем загрузить шрифты в порядке приоритета
    # Сначала Arial (для обычных символов), затем шрифты с поддержкой Unicode (для специальных символов)
    symbol_font = None
    font_name = "default"
    fonts_to_try = []
    
    if platform.system() == "Windows":
        # Список шрифтов в порядке приоритета
        fonts_to_try = [
            ("C:/Windows/Fonts/arial.ttf", "Arial"),           # Arial (для обычных символов)
            ("C:/Windows/Fonts/seguiemj.ttf", "Segoe UI Emoji"), # Лучшая поддержка Unicode
            ("C:/Windows/Fonts/segoeui.ttf", "Segoe UI"),      # Хорошая поддержка Unicode
            ("C:/Windows/Fonts/arialbd.ttf", "Arial Bold"),    # Arial Bold
            ("C:/Windows/Fonts/helvetica.ttf", "Helvetica"),   # Helvetica (если есть)
        ]
    
    # Пробуем загрузить шрифты по очереди
    for font_path, name in fonts_to_try:
        try:
            if os.path.exists(font_path):
                test_font = ImageFont.truetype(font_path, font_size)
                # Проверяем, поддерживает ли шрифт этот символ
                # Рисуем символ на тестовом изображении и проверяем, есть ли контент
                test_img = Image.new("RGBA", (img_size, img_size), (0, 0, 0, 0))
                test_draw = ImageDraw.Draw(test_img)
                test_bbox = test_draw.textbbox((0, 0), symbol, font=test_font)
                test_width = test_bbox[2] - test_bbox[0]
                test_height = test_bbox[3] - test_bbox[1]
                
                # Если размер больше 0, шрифт поддерживает символ
                if test_width > 0 and test_height > 0:
                    test_draw.text((img_size//2, img_size//2), symbol, fill=(255, 255, 255, 255), font=test_font)
                    # Проверяем, есть ли непрозрачные пиксели
                    alpha = test_img.split()[3]
                    has_content = any(alpha.getpixel((i, j)) > 10 for i in range(min(50, test_img.width)) for j in range(min(50, test_img.height)))
                    
                    if has_content:
                        symbol_font = test_font
                        font_name = name
                        break
        except:
            continue
    
    # Если не удалось загрузить системный шрифт, используем дефолтный
    if symbol_font is None:
        symbol_font = ImageFont.load_default()
        if not hasattr(create_symbol_image, '_font_logged'):
            print(f"[PALETTE] Используется дефолтный шрифт для символов (размер: {font_size})")
            create_symbol_image._font_logged = True
    else:
        # Выводим информацию о шрифте только один раз (при первом вызове)
        if not hasattr(create_symbol_image, '_font_logged'):
            print(f"[PALETTE] Используется шрифт {font_name} для символов в палитре (размер: {font_size})")
            create_symbol_image._font_logged = True
    
    # Получаем размер текста
    try:
        bbox = symbol_draw.textbbox((0, 0), symbol, font=symbol_font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
    except:
        # Если не удалось, используем приблизительный размер
        text_width = font_size
        text_height = font_size
    
    # Центрируем символ
    symbol_x = (img_size - text_width) // 2
    symbol_y = (img_size - text_height) // 2
    
    # Рисуем символ
    drawn = False
    try:
        symbol_draw.text((symbol_x, symbol_y), symbol, fill=text_color, font=symbol_font)
        drawn = True
    except:
        # Если не удалось, пробуем белый с обводкой
        try:
            symbol_draw.text((symbol_x, symbol_y), symbol, fill="white", font=symbol_font, stroke_width=1, stroke_fill="black")
            drawn = True
        except:
            try:
                symbol_draw.text((symbol_x, symbol_y), symbol, fill="white", font=symbol_font)
                drawn = True
            except:
                pass
    
    # Проверяем, что символ действительно нарисован (проверяем альфа-канал)
    has_content = False
    if drawn:
        alpha_channel = symbol_img.split()[3]
        # Проверяем наличие непрозрачных пикселей
        for i in range(min(100, symbol_img.width)):
            for j in range(min(100, symbol_img.height)):
                if alpha_channel.getpixel((i, j)) > 10:  # Порог прозрачности
                    has_content = True
                    break
            if has_content:
                break
    
    # Если символ не отображается, выводим предупреждение
    if not has_content:
        symbol_code = ord(symbol[0]) if symbol and len(symbol) > 0 else 'N/A'
        print(f"[WARNING] Символ '{symbol}' (код: {symbol_code}) не отображается шрифтом {font_name}")
    
    # Обрезаем изображение до реального размера символа (с небольшим отступом)
    padding = 5
    if has_content and text_width > 0 and text_height > 0:
        crop_box = (
            max(0, symbol_x - padding),
            max(0, symbol_y - padding),
            min(img_size, symbol_x + text_width + padding),
            min(img_size, symbol_y + text_height + padding)
        )
        if crop_box[2] > crop_box[0] and crop_box[3] > crop_box[1]:
            symbol_img = symbol_img.crop(crop_box)
        else:
            # Если crop_box невалиден, используем минимальный размер
            symbol_img = symbol_img.crop((0, 0, max(font_size, text_width + padding*2), max(font_size, text_height + padding*2)))
    else:
        # Если нет контента, возвращаем минимальное изображение
        symbol_img = Image.new("RGBA", (font_size, font_size), (0, 0, 0, 0))
    
    return symbol_img


def create_palette_image(palette, color_to_char, output_path, numbers=None):
    """
    Создает изображение с визуализацией палитры цветов и символами в цветных квадратиках
    Сетка 10 символов в строке, нумерация сверху
    numbers - список номеров N из Excel (если None, используется idx+1)
    
    НОВАЯ ЛОГИКА: Создает изображения символов заранее и кэширует их для повторного использования
    """
    import math
    
    # Проверяем, что все цвета из палитры есть в color_to_char
    missing_colors = []
    for idx, color in enumerate(palette):
        if color not in color_to_char:
            missing_colors.append((idx, color))
    if missing_colors:
        print(f"[WARNING] Найдено {len(missing_colors)} цветов без символов: {missing_colors[:5]}...")
    
    # Параметры палитры
    cell_size = 50  # Размер ячейки с символом
    cell_padding = 10  # Отступ между ячейками
    symbols_per_row = 10  # Максимум 10 символов в строке
    num_colors = len(palette)
    num_rows = math.ceil(num_colors / symbols_per_row)
    
    # Ширина изображения всегда на основе symbols_per_row (10 колонок)
    # чтобы избежать пустых ячеек или черных квадратов
    actual_columns = symbols_per_row  # Всегда 10 колонок для единообразия
    
    # Высота одной строки: место для нумерации + символы + отступ
    row_height = 40 + cell_size + cell_padding  # 40px для нумерации, cell_size для символа, padding между строками
    
    # Размеры изображения - цифры сверху, картинки снизу
    # Ширина всегда на основе symbols_per_row, чтобы не было пустых мест
    # Минимальные отступы для уменьшения белых полей при вставке в a4_page.jpg
    image_width = actual_columns * (cell_size + cell_padding) + cell_padding
    image_height = num_rows * row_height + 2  # Высота: строки с учетом нумерации + минимальный отступ сверху (уменьшено с 20 до 2)
    
    # Создаем изображение с белым фоном
    palette_img = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(palette_img)
    
    # Загружаем дефолтный шрифт для символов (не из static/fonts)
    symbol_font_size = 24  # Увеличиваем размер шрифта для лучшей видимости
    symbol_font = ImageFont.load_default()
    # Информация о шрифте будет выведена при первом вызове create_symbol_image
    
    # Загружаем шрифт для цифр - делаем его больше и жирнее
    number_font_size = 30
    try:
        number_font = get_font(number_font_size)
    except:
        number_font = ImageFont.load_default()
    
    # Создаем ячейки для каждого цвета
    for idx, color in enumerate(palette):
        # Вычисляем позицию ячейки - переносим на новую строку после 10 символов
        row_idx = idx // symbols_per_row  # Номер строки (0, 1, 2, ...)
        col_idx = idx % symbols_per_row   # Номер колонки в строке (0-9)
        
        x = col_idx * (cell_size + cell_padding) + cell_padding  # Горизонтальное расположение
        y_cards = row_idx * row_height + 40 + 1  # Вертикальное расположение: отступ для нумерации + минимальный промежуток (уменьшено с 10 до 1)
        
        # Получаем символ для этого цвета
        # Нормализуем цвет до кортежа для правильного поиска в словаре
        if isinstance(color, (list, np.ndarray)):
            normalized_color = tuple(int(c) for c in color)
        elif isinstance(color, tuple):
            normalized_color = tuple(int(c) for c in color)
        else:
            normalized_color = color
        
        # Получаем символ из color_to_char (который должен быть загружен из Excel)
        # Символы берутся из столбца "Символ" Excel файла <артикул>.xlsx
        symbol = color_to_char.get(normalized_color, '')
        
        # Очищаем символ от пробелов и проверяем кодировку
        if symbol:
            # Преобразуем в строку, если это не строка
            if not isinstance(symbol, str):
                symbol = str(symbol)
            symbol = symbol.strip()
            
            # Убеждаемся, что символ правильно декодирован (Unicode)
            try:
                # Проверяем, что символ можно закодировать в UTF-8
                symbol.encode('utf-8')
            except (UnicodeEncodeError, AttributeError) as e:
                if idx < 5:
                    print(f"[ERROR] Проблема с кодировкой символа для цвета {normalized_color}: {e}")
                symbol = ''  # Очищаем невалидный символ
        
        # Отладочная информация для всех цветов (чтобы видеть, какие символы есть, а каких нет)
        symbol_repr = repr(symbol) if symbol else "''"
        symbol_code = ord(symbol[0]) if symbol and len(symbol) > 0 else 'N/A'
        
        # Если символ не найден в Excel, выводим предупреждение для ВСЕХ цветов
        if not symbol:
            print(f"[WARNING] Цвет {idx+1}: RGB{normalized_color} - СИМВОЛ НЕ НАЙДЕН в Excel файле!")
        elif idx < 10:  # Для первых 10 выводим подробную информацию
            print(f"[PALETTE] Цвет {idx+1}: RGB{normalized_color}, символ: '{symbol}' (repr: {symbol_repr}, длина: {len(symbol) if symbol else 0}, код: {symbol_code})")
        
        # Рисуем квадратик с цветом из палитры
        draw.rectangle([
            (x, y_cards),
            (x + cell_size, y_cards + cell_size)
        ], fill=color, outline="black", width=2)
        
        # НОВАЯ ЛОГИКА: Рисуем символ внутри квадратика используя функцию create_symbol_image
        if symbol:
            try:
                # Проверяем, что символ не пустой после обработки
                if not symbol or len(symbol) == 0:
                    if idx < 5:
                        print(f"[WARNING] Пустой символ после обработки для цвета {normalized_color} (индекс {idx+1})")
                    continue
                
                # Определяем цвет текста в зависимости от яркости фона
                r, g, b = normalized_color
                luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255.0
                
                # Если фон темный (яркость < 0.5), используем белый текст, иначе черный
                if luminance < 0.5:
                    text_color = "white"
                else:
                    text_color = "black"
                
                # Создаем изображение символа с прозрачным фоном используя новую функцию
                symbol_img = create_symbol_image(symbol, symbol_font_size, text_color=text_color)
                
                # Масштабируем изображение символа до размера ячейки (с небольшим отступом)
                target_size = cell_size - 4  # Оставляем небольшой отступ
                if symbol_img.width > target_size or symbol_img.height > target_size:
                    scale = min(target_size / symbol_img.width, target_size / symbol_img.height)
                    new_size = (int(symbol_img.width * scale), int(symbol_img.height * scale))
                    symbol_img = symbol_img.resize(new_size, Image.Resampling.NEAREST)
                
                # Вычисляем позицию для вставки символа (центрируем в квадратике)
                paste_x = x + (cell_size - symbol_img.width) // 2
                paste_y = y_cards + (cell_size - symbol_img.height) // 2
                
                # Вставляем изображение символа в палитру
                palette_img.paste(symbol_img, (paste_x, paste_y), symbol_img)  # Используем альфа-канал для прозрачности
                
                # Отладочная информация для всех символов
                symbol_code = ord(symbol[0]) if symbol and len(symbol) > 0 else 'N/A'
                if idx < 10:  # Для первых 10 выводим подробную информацию
                    print(f"[PALETTE] Символ {idx+1}: '{symbol}' (код: {symbol_code}) вставлен, размер: {symbol_img.width}x{symbol_img.height}, цвет текста: {text_color}")
                elif idx < 25:  # Для остальных выводим краткую информацию
                    print(f"[PALETTE] Символ {idx+1}: '{symbol}' (код: {symbol_code}) вставлен")
                    
            except Exception as e:
                print(f"[ERROR] Не удалось создать изображение символа '{symbol}' для цвета {normalized_color} (индекс {idx+1}): {e}")
                import traceback
                if idx < 3:
                    traceback.print_exc()
        
        # Добавляем нумерацию сверху от ячейки (используем N из Excel, если доступно)
        if numbers and idx < len(numbers):
            number_text = str(numbers[idx])  # Номер из Excel (N)
        else:
            number_text = str(idx + 1)  # Нумерация начинается с 1
        try:
            bbox_num = draw.textbbox((0, 0), number_text, font=number_font)
            num_width = bbox_num[2] - bbox_num[0]
            num_height = bbox_num[3] - bbox_num[1]
        except:
            num_width = number_font_size
            num_height = number_font_size
        
        num_x = x + (cell_size - num_width) // 2  # По центру по горизонтали
        num_y = row_idx * row_height + 5  # Цифры в начале строки с небольшим отступом сверху
        
        # Рисуем цифру с черной обводкой для лучшей видимости
        try:
            draw.text((num_x, num_y), number_text, fill="black", font=number_font, stroke_width=1, stroke_fill="white")
        except:
            # Если stroke не поддерживается, рисуем без обводки
            draw.text((num_x, num_y), number_text, fill="black", font=number_font)
    
    # Итоговая сводка по символам
    symbols_with_content = sum(1 for idx, color in enumerate(palette) if color in color_to_char and color_to_char[color])
    symbols_missing = num_colors - symbols_with_content
    print(f"[PALETTE] Итоговая сводка: {symbols_with_content}/{num_colors} символов успешно добавлено в палитру")
    if symbols_missing > 0:
        print(f"[PALETTE] ВНИМАНИЕ: {symbols_missing} цветов без символов!")
    
    # Сохраняем изображение в формате JPG с метаданными
    palette_img_rgb = palette_img.convert("RGB")
    save_jpg_with_metadata(palette_img_rgb, output_path, quality=config.JPEG_QUALITY, dpi=(config.DPI, config.DPI), title="Палитра цветов")


def create_palette_pdf(palette, color_to_char, output_pdf_path, numbers=None):
    """
    Создает PDF с визуализацией палитры цветов и символами в цветных квадратиках
    Сетка 10 символов в строке, нумерация сверху
    numbers - список номеров N из Excel (если None, используется idx+1)
    """
    import math
    
    # Проверяем, что все цвета из палитры есть в color_to_char
    missing_colors = []
    for idx, color in enumerate(palette):
        if color not in color_to_char:
            missing_colors.append((idx, color))
    if missing_colors:
        print(f"[WARNING] Найдено {len(missing_colors)} цветов без символов: {missing_colors[:5]}...")
    
    # Параметры палитры (в мм для PDF)
    cell_size_mm = 10  # Размер ячейки с символом в мм
    cell_padding_mm = 2  # Отступ между ячейками в мм
    symbols_per_row = 10  # Максимум 10 символов в строке
    num_colors = len(palette)
    num_rows = math.ceil(num_colors / symbols_per_row)
    
    # Высота одной строки: место для нумерации + символы + отступ
    number_height_mm = 8  # Высота для нумерации
    row_height_mm = number_height_mm + cell_size_mm + cell_padding_mm
    
    # Размеры страницы
    # Минимальные отступы для уменьшения белых полей при вставке в a4_page.jpg
    page_margin_mm = 1  # Уменьшено с 10 до 1 мм для минимальных отступов
    page_width = (symbols_per_row * (cell_size_mm + cell_padding_mm) + cell_padding_mm + 2 * page_margin_mm) * mm
    page_height = (num_rows * row_height_mm + 2 * page_margin_mm) * mm
    
    # Создаем PDF
    c = canvas.Canvas(output_pdf_path, pagesize=(page_width, page_height))
    set_pdf_metadata(c, "Палитра цветов")
    
    # Параметры шрифтов
    symbol_font_size = 18  # Размер шрифта для символов
    number_font_size = 8  # Размер шрифта для нумерации
    
    # Создаем ячейки для каждого цвета
    for idx, color in enumerate(palette):
        # Вычисляем позицию ячейки
        row_idx = idx // symbols_per_row
        col_idx = idx % symbols_per_row
        
        x = (col_idx * (cell_size_mm + cell_padding_mm) + cell_padding_mm + page_margin_mm) * mm
        # В reportlab координаты идут снизу вверх
        # Вычисляем y для нижнего края ячейки (y_cards - это нижний левый угол прямоугольника)
        # Для первой строки (row_idx=0): y = page_height - page_margin - number_height - cell_size
        y_cards = page_height - (page_margin_mm + number_height_mm + cell_size_mm + row_idx * row_height_mm) * mm
        
        # Нормализуем цвет
        if isinstance(color, (list, np.ndarray)):
            normalized_color = tuple(int(c) for c in color)
        elif isinstance(color, tuple):
            normalized_color = tuple(int(c) for c in color)
        else:
            normalized_color = color
        
        # Получаем символ
        symbol = color_to_char.get(normalized_color, '')
        if symbol:
            if not isinstance(symbol, str):
                symbol = str(symbol)
            symbol = symbol.strip()
            try:
                symbol.encode('utf-8')
            except (UnicodeEncodeError, AttributeError):
                symbol = ''
        
        # Рисуем квадратик с цветом
        r, g, b = normalized_color
        c.setFillColorRGB(r/255.0, g/255.0, b/255.0)
        c.setStrokeColorRGB(0, 0, 0)
        c.setLineWidth(0.5)
        c.rect(x, y_cards, cell_size_mm * mm, cell_size_mm * mm, fill=1, stroke=1)
        
        # Рисуем символ внутри квадратика
        if symbol:
            # Определяем цвет текста в зависимости от яркости фона
            luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255.0
            if luminance < 0.5:
                text_color = (1, 1, 1)  # Белый
            else:
                text_color = (0, 0, 0)  # Черный
            
            c.setFillColorRGB(*text_color)
            c.setFont("Helvetica", symbol_font_size)
            
            # Центрируем символ в ячейке
            text_x = x + (cell_size_mm * mm) / 2
            text_y = y_cards + (cell_size_mm * mm) / 2 - symbol_font_size / 3
            c.drawCentredString(text_x, text_y, symbol)
        
        # Добавляем нумерацию сверху
        if numbers and idx < len(numbers):
            number_text = str(numbers[idx])
        else:
            number_text = str(idx + 1)
        
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", number_font_size)
        num_x = x + (cell_size_mm * mm) / 2
        # Нумерация сверху от ячейки (в reportlab координаты снизу вверх)
        # y_cards - это нижний край ячейки, поэтому нумерация выше на cell_size + небольшой отступ
        num_y = y_cards + cell_size_mm * mm + 1 * mm
        c.drawCentredString(num_x, num_y, number_text)
    
    # Итоговая сводка
    symbols_with_content = sum(1 for idx, color in enumerate(palette) if color in color_to_char and color_to_char[color])
    symbols_missing = num_colors - symbols_with_content
    print(f"[PALETTE-PDF] Итоговая сводка: {symbols_with_content}/{num_colors} символов успешно добавлено в палитру")
    if symbols_missing > 0:
        print(f"[PALETTE-PDF] ВНИМАНИЕ: {symbols_missing} цветов без символов!")
    
    # Копирайт не добавляем для layout_palette.pdf
    
    c.save()
    print(f"[PALETTE-PDF] PDF палитры сохранен: {output_pdf_path}")


def pdf_to_jpg(pdf_path, jpg_path, dpi=None):
    """
    Конвертирует PDF в JPG изображение
    Использует PyMuPDF (fitz) если доступен, иначе pdf2image
    """
    if dpi is None:
        dpi = config.DPI
    """
    Конвертирует PDF в JPG изображение
    Использует PyMuPDF (fitz) если доступен, иначе pdf2image
    """
    try:
        # Пробуем использовать PyMuPDF (fitz)
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(pdf_path)
            if len(doc) > 0:
                page = doc[0]
                # Конвертируем в изображение с заданным DPI
                zoom = dpi / 72.0
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat)
                # Конвертируем в PIL Image и сохраняем как JPG с метаданными
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                save_jpg_with_metadata(img, jpg_path, quality=config.JPEG_QUALITY, dpi=(dpi, dpi), title="Схема для вышивания")
                doc.close()
                print(f"[PDF2JPG] PDF конвертирован в JPG с помощью PyMuPDF: {jpg_path}")
                return True
        except ImportError:
            pass
        
        # Пробуем использовать pdf2image
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(pdf_path, dpi=dpi)
            if images:
                # Конвертируем в RGB и сохраняем как JPG с метаданными
                img_rgb = images[0].convert("RGB")
                save_jpg_with_metadata(img_rgb, jpg_path, quality=config.JPEG_QUALITY, dpi=(dpi, dpi), title="Схема для вышивания")
                print(f"[PDF2JPG] PDF конвертирован в JPG с помощью pdf2image: {jpg_path}")
                return True
        except ImportError:
            pass
        
        print(f"[ERROR] Не удалось конвертировать PDF в JPG. Установите PyMuPDF (pip install PyMuPDF) или pdf2image (pip install pdf2image)")
        return False
        
    except Exception as e:
        print(f"[ERROR] Ошибка при конвертации PDF в JPG: {e}")
        import traceback
        traceback.print_exc()
        return False


def create_pdf_table(image_path, output_pdf_path, num_blocks_width=None, num_blocks_height=None, include_symbols=False, palette=None, painted_cells=None, vertical_lines=None, horizontal_lines=None, brand='marfa'):
    """
    Создает PDF с таблицей, где каждая клетка покрашена в цвет из изображения
    Похоже на структуру из KoiFishesWaterLilyPatternKeeper.pdf
    
    Args:
        painted_cells: dict - словарь {(col, row): color} с закрашенными ячейками
        vertical_lines: list - вертикальные линии сетки для определения индексов ячеек
        horizontal_lines: list - горизонтальные линии сетки для определения индексов ячеек
    """
    if num_blocks_width is None:
        num_blocks_width = config.NUM_BLOCKS_WIDTH
    if num_blocks_height is None:
        num_blocks_height = config.NUM_BLOCKS_HEIGHT
    """
    Создает PDF с таблицей, где каждая клетка покрашена в цвет из изображения
    Похоже на структуру из KoiFishesWaterLilyPatternKeeper.pdf
    """
    # Загружаем изображение
    img = Image.open(image_path)
    img = img.convert("RGB")
    img_width, img_height = img.size
    
    # Вычисляем размер клетки на основе размера изображения
    cell_width = img_width / num_blocks_width
    cell_height = img_height / num_blocks_height
    
    # Используем большой размер страницы для таблицы 150x150
    # Создаем кастомный размер страницы, чтобы таблица поместилась
    # Для 150x150 с клетками минимум 2mm каждая нужно минимум 300mm
    # Используем несколько страниц или один большой формат
    page_width = 800 * mm  # Большой формат
    page_height = 800 * mm
    
    c = canvas.Canvas(output_pdf_path, pagesize=(page_width, page_height))
    set_pdf_metadata(c, "Таблица схемы для вышивания")
    
    # Отступы
    margin = 10 * mm
    
    # Вычисляем размер клетки в PDF (в точках)
    # Минимум из конфига на клетку для читаемости
    min_cell_size = config.MIN_CELL_SIZE_MM * mm
    available_width = page_width - 2 * margin
    available_height = page_height - 2 * margin
    
    cell_size_pt = max(min_cell_size, min(available_width / num_blocks_width, available_height / num_blocks_height))
    
    # Вычисляем общий размер таблицы
    table_width = cell_size_pt * num_blocks_width
    table_height = cell_size_pt * num_blocks_height
    
    # Центрируем таблицу на странице
    start_x = (page_width - table_width) / 2
    start_y = (page_height - table_height) / 2
    
    # Используем переданную палитру или извлекаем из изображения
    if palette is None:
        palette = extract_dominant_colors(image_path, 25)
    char_list = SPECIAL_CHARS * ((len(palette) // len(SPECIAL_CHARS)) + 1)
    color_to_char = {color: char_list[i] for i, color in enumerate(palette)}
    
    # Функция для определения индексов ячейки по координатам пикселя
    def get_cell_indices(img_x, img_y):
        """Определяет индексы ячейки (col, row) по координатам пикселя"""
        if not vertical_lines or not horizontal_lines:
            return None, None
        col = -1
        row = -1
        for i in range(len(vertical_lines) - 1):
            if vertical_lines[i] <= img_x < vertical_lines[i + 1]:
                col = i
                break
        for i in range(len(horizontal_lines) - 1):
            if horizontal_lines[i] <= img_y < horizontal_lines[i + 1]:
                row = i
                break
        if col >= 0 and row >= 0:
            return (col, row)
        return None, None
    
    # Белый фон для ячеек без символов (не в painted_cells)
    LIGHT_BEIGE = (255, 255, 255)  # белый фон для ячеек без символов
    
    # Рисуем таблицу с сеткой
    for y in range(num_blocks_height):
        for x in range(num_blocks_width):
            # Получаем цвет из изображения (layout.png уже содержит цвета из палитры)
            pixel_x = int(x * cell_width + cell_width / 2)
            pixel_y = int(y * cell_height + cell_height / 2)
            
            # Ограничиваем координаты
            pixel_x = min(pixel_x, img_width - 1)
            pixel_y = min(pixel_y, img_height - 1)
            
            color = img.getpixel((pixel_x, pixel_y))
            
            # Определяем индексы ячейки для проверки, закрашена ли она
            col, row = get_cell_indices(pixel_x, pixel_y)
            is_painted = (col, row) in painted_cells if (painted_cells and col is not None and row is not None) else True
            
            # Находим ближайший цвет из палитры (используем ту же палитру, что и для layout)
            closest_color = min(palette, key=lambda c: color_distance(color, c))
            
            # Вычисляем позицию клетки в PDF
            pdf_x = start_x + x * cell_size_pt
            pdf_y = start_y + (num_blocks_height - 1 - y) * cell_size_pt  # Инвертируем Y
            
            # Ячейки без символов (не в painted_cells) — белый фон; остальные — цвет из палитры
            if is_painted:
                fill_color = closest_color
            else:
                fill_color = LIGHT_BEIGE
            c.setFillColorRGB(fill_color[0]/255.0, fill_color[1]/255.0, fill_color[2]/255.0)
            c.setStrokeColorRGB(0, 0, 0)  # Черная рамка
            c.setLineWidth(0.1)  # Тонкая линия для всех границ клетки
            c.rect(pdf_x, pdf_y, cell_size_pt, cell_size_pt, fill=1, stroke=1)
            
            # Если нужно добавить символы и размер клетки достаточен
            if include_symbols and cell_size_pt > 1.5 * mm:
                # Получаем символ только если ячейка закрашена (есть в painted_cells)
                # Если ячейка стерта (нет в painted_cells), символ не присваиваем
                if is_painted:
                    symbol = color_to_char.get(closest_color, '')
                else:
                    symbol = ''  # Стертая ячейка - символ не нужен
                
                # Рисуем символ, если он есть
                if symbol:
                    try:
                        font_size = max(config.MIN_SYMBOL_SIZE_PT, cell_size_pt * config.SYMBOL_SIZE_RATIO)
                        # Центрируем символ в клетке
                        text_x = pdf_x + cell_size_pt / 2
                        text_y = pdf_y + cell_size_pt / 2 - font_size / 3
                        
                        # Выбираем оптимальный цвет текста для читаемости (принцип из Pixel Logic)
                        # Всегда используем контрастный цвет для лучшей видимости символов
                        text_color = get_optimal_text_color(closest_color)
                        c.setFillColorRGB(text_color[0]/255.0, text_color[1]/255.0, text_color[2]/255.0)
                        
                        # Используем только базовый шрифт Helvetica
                        c.setFont("Helvetica", font_size)
                        c.drawCentredString(text_x, text_y, symbol)
                    except Exception as e:
                        # Если символ не поддерживается, пробуем альтернативный способ с TextObject
                        try:
                            font_size = max(config.MIN_SYMBOL_SIZE_PT, cell_size_pt * config.SYMBOL_SIZE_RATIO)
                            text_x = pdf_x + cell_size_pt / 2
                            text_y = pdf_y + cell_size_pt / 2
                            
                            # Используем TextObject для точного контроля цвета
                            text_obj = c.beginText(text_x, text_y)
                            # Используем только базовый шрифт Helvetica
                            text_obj.setFont("Helvetica", font_size)
                            # Выбираем оптимальный цвет текста для читаемости (принцип из Pixel Logic)
                            # Всегда используем контрастный цвет для лучшей видимости символов
                            text_color = get_optimal_text_color(closest_color)
                            text_obj.setFillColorRGB(text_color[0]/255.0, text_color[1]/255.0, text_color[2]/255.0)
                            # Используем moveCursor для центрирования
                            bbox = c.stringWidth(symbol, "Helvetica", font_size)
                            text_obj.moveCursor(-bbox/2, -font_size/3)
                            text_obj.textOut(symbol)
                            c.drawText(text_obj)
                        except:
                            pass
        
    # Рисуем жирные линии деления по 10 (только на нужных границах)
    print(f"[PDF] Добавление жирных линий деления по 10...")
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(2.0)  # Жирная линия
    
    # Вертикальные линии (каждые 10 столбцов) - только ЛЕВАЯ граница столбца
    # Рисуем линию слева от столбцов 10, 20, 30 и т.д.
    for x in range(10, num_blocks_width + 1, 10):
        line_x = start_x + x * cell_size_pt
        c.line(line_x, start_y, line_x, start_y + table_height)
    
    # Горизонтальные линии (каждые 10 строк) - только ВЕРХНЯЯ граница строки
    # Рисуем линию сверху от строк 10, 20, 30 и т.д.
    for y in range(10, num_blocks_height + 1, 10):
        # Инвертируем Y для правильного позиционирования
        line_y = start_y + (num_blocks_height - y) * cell_size_pt
        c.line(start_x, line_y, start_x + table_width, line_y)
    
    # Рисуем красные линии центра (вертикальная и горизонтальная)
    print(f"[PDF] Добавление красных линий центра...")
    center_x = num_blocks_width // 2  # Центр по X (75 для 150)
    center_y = num_blocks_height // 2  # Центр по Y (75 для 150)
    
    # Устанавливаем параметры для красных линий
    c.setStrokeColorRGB(1, 0, 0)  # Красный цвет
    c.setLineWidth(2.0)  # Толщина линии для центра
    
    # Вертикальная красная линия через центр
    center_line_x = start_x + center_x * cell_size_pt
    c.line(center_line_x, start_y, center_line_x, start_y + table_height)
    
    # Горизонтальная красная линия через центр
    center_line_y = start_y + (num_blocks_height - center_y) * cell_size_pt
    c.line(start_x, center_line_y, start_x + table_width, center_line_y)
    
    # Добавляем номера по оси X (сверху) - ПЕРЕД стрелками
    print(f"[PDF] Добавление номеров по оси X...")
    c.setFillColorRGB(0, 0, 0)
    scale_font_name, scale_font_size = get_scale_font(config.SCALE_FONT_SIZE)
    c.setFont(scale_font_name, scale_font_size)
    number_offset_y = 3 * mm  # Отступ для номеров сверху
    
    for x in range(0, num_blocks_width + 1, 10):
        number_x = start_x + x * cell_size_pt
        number_text = str(x)
        text_width = c.stringWidth(number_text, scale_font_name, scale_font_size)
        c.drawCentredString(number_x, start_y + table_height + number_offset_y, number_text)
    
    # Добавляем номера по оси Y (слева) - ПЕРЕД стрелками
    print(f"[PDF] Добавление номеров по оси Y...")
    number_offset_x = 3 * mm  # Отступ для номеров слева
    
    for y in range(10, num_blocks_height + 1, 10):
        # Инвертируем Y для правильного отображения
        pdf_y = start_y + (num_blocks_height - y) * cell_size_pt
        number_text = str(y)
        text_width = c.stringWidth(number_text, scale_font_name, scale_font_size)
        c.drawRightString(start_x - number_offset_x, pdf_y - 4, number_text)
    
    # Рисуем стрелочки на границах - ПОСЛЕ номеров
    print(f"[PDF] Добавление стрелочек на границах...")
    arrow_size = 8 * mm  # Размер стрелочек
    arrow_line_width = 2.0
    arrow_offset = 10 * mm  # Отступ стрелок от границ таблицы
    arrow_head_size = 4 * mm
    
    # Вычисляем позиции стрелок
    arrow_top_x = center_line_x
    arrow_top_y = start_y + table_height + arrow_offset
    arrow_bottom_x = center_line_x
    arrow_bottom_y = start_y - arrow_offset
    arrow_left_x = start_x - arrow_offset
    arrow_left_y = center_line_y
    arrow_right_x = start_x + table_width + arrow_offset
    arrow_right_y = center_line_y
    
    # Рисуем белые прямоугольники под стрелками для перекрытия номеров
    c.setFillColorRGB(1, 1, 1)  # Белый цвет
    c.setStrokeColorRGB(1, 1, 1)
    
    # Увеличиваем размеры белого прямоугольника
    white_box_size_horizontal = 12 * mm  # Ширина для горизонтальной стрелки
    white_box_size_vertical = 8 * mm  # Высота для вертикальной стрелки
    white_box_padding = 3 * mm  # Запас вокруг
    
    # Вычисляем размер текста для самого большого номера
    max_number = max(num_blocks_width, num_blocks_height)
    max_number_text = str(max_number)
    max_text_width = c.stringWidth(max_number_text, scale_font_name, scale_font_size)
    text_height = scale_font_size * 1.3
    
    # Используем максимум из фиксированного размера и размера текста
    white_box_width = max(white_box_size_horizontal, max_text_width + 4 * mm)
    white_box_height = max(white_box_size_vertical, text_height + 2 * mm)
    
    # Белый прямоугольник под стрелкой сверху (перекрывает номер, если он попадает на центр)
    c.rect(arrow_top_x - white_box_width/2, 
           start_y + table_height + number_offset_y - white_box_padding,
           white_box_width, 
           white_box_height, 
           fill=1, stroke=0)
    
    # Белый прямоугольник под стрелкой слева (перекрывает номер, если он попадает на центр)
    c.rect(start_x - number_offset_x - white_box_height - 2 * mm,
           arrow_left_y - white_box_width/2,
           white_box_height + 2 * mm,
           white_box_width,
           fill=1, stroke=0)
    
    # Теперь рисуем стрелки поверх белых прямоугольников
    c.setStrokeColorRGB(0, 0, 0)
    c.setFillColorRGB(0, 0, 0)
    c.setLineWidth(arrow_line_width)
    
    # Стрелочка вверху (центр по X) - указывает вниз к центру (↓)
    # Вертикальная линия стрелки (вниз к центру)
    c.line(arrow_top_x, arrow_top_y, arrow_top_x, arrow_top_y - arrow_size)
    # Голова стрелки (треугольник, указывает вниз)
    c.line(arrow_top_x, arrow_top_y - arrow_size, arrow_top_x - arrow_head_size/2, arrow_top_y - arrow_size + arrow_head_size)
    c.line(arrow_top_x, arrow_top_y - arrow_size, arrow_top_x + arrow_head_size/2, arrow_top_y - arrow_size + arrow_head_size)
    
    # Стрелочка внизу (центр по X) - указывает вверх к центру (↑)
    # Рисуем стрелку вверх (к центру)
    c.line(arrow_bottom_x, arrow_bottom_y, arrow_bottom_x, arrow_bottom_y + arrow_size)
    # Голова стрелки (треугольник, указывает вверх)
    c.line(arrow_bottom_x, arrow_bottom_y + arrow_size, arrow_bottom_x - arrow_head_size/2, arrow_bottom_y + arrow_size - arrow_head_size)
    c.line(arrow_bottom_x, arrow_bottom_y + arrow_size, arrow_bottom_x + arrow_head_size/2, arrow_bottom_y + arrow_size - arrow_head_size)
    
    # Стрелочка слева (центр по Y) - указывает вправо к центру (→)
    # Рисуем стрелку вправо (к центру)
    c.line(arrow_left_x, arrow_left_y, arrow_left_x + arrow_size, arrow_left_y)
    # Голова стрелки (треугольник, указывает вправо)
    c.line(arrow_left_x + arrow_size, arrow_left_y, arrow_left_x + arrow_size - arrow_head_size, arrow_left_y - arrow_head_size/2)
    c.line(arrow_left_x + arrow_size, arrow_left_y, arrow_left_x + arrow_size - arrow_head_size, arrow_left_y + arrow_head_size/2)
    
    # Стрелочка справа (центр по Y) - указывает влево к центру (←)
    # Рисуем стрелку влево (к центру)
    c.line(arrow_right_x, arrow_right_y, arrow_right_x - arrow_size, arrow_right_y)
    # Голова стрелки (треугольник, указывает влево)
    c.line(arrow_right_x - arrow_size, arrow_right_y, arrow_right_x - arrow_size + arrow_head_size, arrow_right_y - arrow_head_size/2)
    c.line(arrow_right_x - arrow_size, arrow_right_y, arrow_right_x - arrow_size + arrow_head_size, arrow_right_y + arrow_head_size/2)
    
    # Возвращаем черный цвет для дальнейшей отрисовки
    c.setStrokeColorRGB(0, 0, 0)
    
    # Добавляем копирайт
    add_copyright_to_pdf(c, brand=brand)
    
    c.save()


def create_pdf_symbols_only(image_path, output_pdf_path, num_blocks_width=None, num_blocks_height=None, palette=None, painted_cells=None, vertical_lines=None, horizontal_lines=None, brand='marfa'):
    """
    Создает PDF только с символами (без цветных блоков)
    
    Args:
        painted_cells: dict - словарь {(col, row): color} с закрашенными ячейками
        vertical_lines: list - вертикальные линии сетки для определения индексов ячеек
        horizontal_lines: list - горизонтальные линии сетки для определения индексов ячеек
    """
    if num_blocks_width is None:
        num_blocks_width = config.NUM_BLOCKS_WIDTH
    if num_blocks_height is None:
        num_blocks_height = config.NUM_BLOCKS_HEIGHT
    """
    Создает PDF с таблицей, где только черные символы без цветных блоков
    """
    # Загружаем изображение
    img = Image.open(image_path)
    img = img.convert("RGB")
    img_width, img_height = img.size
    
    # Вычисляем размер клетки на основе размера изображения
    cell_width = img_width / num_blocks_width
    cell_height = img_height / num_blocks_height
    
    # Используем большой размер страницы для таблицы 150x150
    page_width = 800 * mm
    page_height = 800 * mm
    
    c = canvas.Canvas(output_pdf_path, pagesize=(page_width, page_height))
    set_pdf_metadata(c, "Таблица символов для вышивания")
    
    # Отступы
    margin = 10 * mm
    
    # Вычисляем размер клетки в PDF (в точках)
    min_cell_size = config.MIN_CELL_SIZE_MM * mm
    available_width = page_width - 2 * margin
    available_height = page_height - 2 * margin
    
    cell_size_pt = max(min_cell_size, min(available_width / num_blocks_width, available_height / num_blocks_height))
    
    # Вычисляем общий размер таблицы
    table_width = cell_size_pt * num_blocks_width
    table_height = cell_size_pt * num_blocks_height
    
    # Центрируем таблицу на странице
    start_x = (page_width - table_width) / 2
    start_y = (page_height - table_height) / 2
    
    # Используем переданную палитру или извлекаем из изображения
    if palette is None:
        palette = extract_dominant_colors(image_path, 25)
    char_list = SPECIAL_CHARS * ((len(palette) // len(SPECIAL_CHARS)) + 1)
    color_to_char = {color: char_list[i] for i, color in enumerate(palette)}
    
    # Функция для определения индексов ячейки по координатам пикселя
    def get_cell_indices(img_x, img_y):
        """Определяет индексы ячейки (col, row) по координатам пикселя"""
        if not vertical_lines or not horizontal_lines:
            return None, None
        col = -1
        row = -1
        for i in range(len(vertical_lines) - 1):
            if vertical_lines[i] <= img_x < vertical_lines[i + 1]:
                col = i
                break
        for i in range(len(horizontal_lines) - 1):
            if horizontal_lines[i] <= img_y < horizontal_lines[i + 1]:
                row = i
                break
        if col >= 0 and row >= 0:
            return (col, row)
        return None, None
    
    # Рисуем только символы (без цветных блоков)
    for y in range(num_blocks_height):
        for x in range(num_blocks_width):
            # Получаем цвет из изображения
            pixel_x = int(x * cell_width + cell_width / 2)
            pixel_y = int(y * cell_height + cell_height / 2)
            
            # Ограничиваем координаты
            pixel_x = min(pixel_x, img_width - 1)
            pixel_y = min(pixel_y, img_height - 1)
            
            color = img.getpixel((pixel_x, pixel_y))
            
            # Определяем индексы ячейки для проверки, закрашена ли она
            col, row = get_cell_indices(pixel_x, pixel_y)
            is_painted = (col, row) in painted_cells if (painted_cells and col is not None and row is not None) else True
            
            # Вычисляем позицию клетки в PDF
            pdf_x = start_x + x * cell_size_pt
            pdf_y = start_y + (num_blocks_height - 1 - y) * cell_size_pt  # Инвертируем Y
            
            # Находим ближайший цвет из палитры для получения символа
            closest_color = min(palette, key=lambda c: color_distance(color, c))
            
            # Получаем символ только если ячейка закрашена (есть в painted_cells)
            # Если ячейка стерта (нет в painted_cells), символ не присваиваем
            if is_painted:
                symbol = color_to_char.get(closest_color, '')
            else:
                symbol = ''  # Стертая ячейка - символ не нужен
            
            # Рисуем только символ (без цветного блока), если размер клетки достаточен
            if symbol and cell_size_pt > 1.5 * mm:
                try:
                    font_size = max(config.MIN_SYMBOL_SIZE_PT, cell_size_pt * config.SYMBOL_SIZE_RATIO)  # Используем те же параметры, что и в create_pdf_table (70% и минимум 8pt)
                    # Центрируем символ в клетке
                    text_x = pdf_x + cell_size_pt / 2
                    text_y = pdf_y + cell_size_pt / 2 - font_size / 3
                    
                    # Выбираем оптимальный цвет текста для читаемости (принцип из Pixel Logic)
                    auto_text_color = getattr(config, 'AUTO_TEXT_COLOR', False)
                    if auto_text_color:
                        text_color = get_optimal_text_color(closest_color)
                        c.setFillColorRGB(text_color[0]/255.0, text_color[1]/255.0, text_color[2]/255.0)
                    else:
                        # Устанавливаем черный цвет для символа (по умолчанию)
                        c.setFillColorRGB(0, 0, 0)  # Черный цвет
                    # Используем только базовый шрифт Helvetica
                    c.setFont("Helvetica", font_size)
                    c.drawCentredString(text_x, text_y, symbol)
                except Exception as e:
                    # Если символ не поддерживается, пробуем альтернативный способ с TextObject
                    try:
                        font_size = max(config.MIN_SYMBOL_SIZE_PT, cell_size_pt * config.SYMBOL_SIZE_RATIO)
                        text_x = pdf_x + cell_size_pt / 2
                        text_y = pdf_y + cell_size_pt / 2
                        
                        # Используем TextObject для точного контроля цвета
                        text_obj = c.beginText(text_x, text_y)
                        # Используем только базовый шрифт Helvetica
                        font_name = "Helvetica"
                        text_obj.setFont("Helvetica", font_size)
                        # Выбираем оптимальный цвет текста для читаемости (принцип из Pixel Logic)
                        auto_text_color = getattr(config, 'AUTO_TEXT_COLOR', False)
                        if auto_text_color:
                            text_color = get_optimal_text_color(closest_color)
                            text_obj.setFillColorRGB(text_color[0]/255.0, text_color[1]/255.0, text_color[2]/255.0)
                        else:
                            text_obj.setFillColorRGB(0, 0, 0)  # Черный цвет (по умолчанию)
                        # Используем moveCursor для центрирования
                        bbox = c.stringWidth(symbol, font_name, font_size)
                        text_obj.moveCursor(-bbox/2, -font_size/3)
                        text_obj.textOut(symbol)
                        c.drawText(text_obj)
                    except:
                        pass
        
    # Рисуем тонкую сетку для всех клеток (без заливки цветом)
    print(f"[PDF] Добавление тонкой сетки...")
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.1)  # Тонкая линия для сетки
    
    # Вертикальные линии для всех столбцов
    for x in range(0, num_blocks_width + 1):
        line_x = start_x + x * cell_size_pt
        c.line(line_x, start_y, line_x, start_y + table_height)
    
    # Горизонтальные линии для всех строк
    for y in range(0, num_blocks_height + 1):
        line_y = start_y + (num_blocks_height - y) * cell_size_pt
        c.line(start_x, line_y, start_x + table_width, line_y)
    
    # Рисуем жирные линии деления по 10
    print(f"[PDF] Добавление жирных линий деления по 10...")
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(2.0)  # Жирная линия
    
    # Вертикальные линии (каждые 10 столбцов)
    for x in range(10, num_blocks_width + 1, 10):
        line_x = start_x + x * cell_size_pt
        c.line(line_x, start_y, line_x, start_y + table_height)
    
    # Горизонтальные линии (каждые 10 строк)
    for y in range(10, num_blocks_height + 1, 10):
        line_y = start_y + (num_blocks_height - y) * cell_size_pt
        c.line(start_x, line_y, start_x + table_width, line_y)
    
    # Рисуем красные линии центра (вертикальная и горизонтальная)
    print(f"[PDF] Добавление красных линий центра...")
    center_x = num_blocks_width // 2  # Центр по X (75 для 150)
    center_y = num_blocks_height // 2  # Центр по Y (75 для 150)
    
    # Устанавливаем параметры для красных линий
    c.setStrokeColorRGB(1, 0, 0)  # Красный цвет
    c.setLineWidth(2.0)  # Толщина линии для центра
    
    # Вертикальная красная линия через центр
    center_line_x = start_x + center_x * cell_size_pt
    c.line(center_line_x, start_y, center_line_x, start_y + table_height)
    
    # Горизонтальная красная линия через центр
    center_line_y = start_y + (num_blocks_height - center_y) * cell_size_pt
    c.line(start_x, center_line_y, start_x + table_width, center_line_y)
    
    # Добавляем номера по оси X (сверху) - ПЕРЕД стрелками
    print(f"[PDF] Добавление номеров по оси X...")
    c.setFillColorRGB(0, 0, 0)
    scale_font_name, scale_font_size = get_scale_font(config.SCALE_FONT_SIZE)
    c.setFont(scale_font_name, scale_font_size)
    number_offset_y = 3 * mm
    
    for x in range(0, num_blocks_width + 1, 10):
        number_x = start_x + x * cell_size_pt
        number_text = str(x)
        c.drawCentredString(number_x, start_y + table_height + number_offset_y, number_text)
    
    # Добавляем номера по оси Y (слева) - ПЕРЕД стрелками
    print(f"[PDF] Добавление номеров по оси Y...")
    number_offset_x = 3 * mm
    
    for y in range(10, num_blocks_height + 1, 10):
        pdf_y = start_y + (num_blocks_height - y) * cell_size_pt
        number_text = str(y)
        c.drawRightString(start_x - number_offset_x, pdf_y - 4, number_text)
    
    # Рисуем стрелочки на границах - ПОСЛЕ номеров
    print(f"[PDF] Добавление стрелочек на границах...")
    arrow_size = 8 * mm  # Размер стрелочек
    arrow_line_width = 2.0
    arrow_offset = 10 * mm  # Отступ стрелок от границ таблицы
    arrow_head_size = 4 * mm
    
    # Вычисляем позиции стрелок
    arrow_top_x = center_line_x
    arrow_top_y = start_y + table_height + arrow_offset
    arrow_bottom_x = center_line_x
    arrow_bottom_y = start_y - arrow_offset
    arrow_left_x = start_x - arrow_offset
    arrow_left_y = center_line_y
    arrow_right_x = start_x + table_width + arrow_offset
    arrow_right_y = center_line_y
    
    # Рисуем белые прямоугольники под стрелками для перекрытия номеров
    c.setFillColorRGB(1, 1, 1)  # Белый цвет
    c.setStrokeColorRGB(1, 1, 1)
    
    # Увеличиваем размеры белого прямоугольника
    white_box_size_horizontal = 12 * mm  # Ширина для горизонтальной стрелки
    white_box_size_vertical = 7 * mm  # Высота для вертикальной стрелки
    white_box_padding = 1 * mm  # Запас вокруг
    
    # Вычисляем размер текста для самого большого номера
    max_number = max(num_blocks_width, num_blocks_height)
    max_number_text = str(max_number)
    max_text_width = c.stringWidth(max_number_text, scale_font_name, scale_font_size)
    text_height = scale_font_size * 1.3
    
    # Используем максимум из фиксированного размера и размера текста
    white_box_width = max(white_box_size_horizontal, max_text_width + 4 * mm)
    white_box_height = max(white_box_size_vertical, text_height + 2 * mm)
    
    # Белый прямоугольник под стрелкой сверху (перекрывает номер, если он попадает на центр)
    c.rect(arrow_top_x - white_box_width/2, 
           start_y + table_height + number_offset_y - white_box_padding,
           white_box_width, 
           white_box_height, 
           fill=1, stroke=0)
    
    # Белый прямоугольник под стрелкой слева (перекрывает номер, если он попадает на центр)
    c.rect(start_x - number_offset_x - white_box_height - 2 * mm,
           arrow_left_y - white_box_width/2,
           white_box_height + 2 * mm,
           white_box_width,
           fill=1, stroke=0)
    
    # Теперь рисуем стрелки поверх белых прямоугольников
    c.setStrokeColorRGB(0, 0, 0)
    c.setFillColorRGB(0, 0, 0)
    c.setLineWidth(arrow_line_width)
    
    # Стрелочка вверху (центр по X) - указывает вниз к центру (↓)
    # Вертикальная линия стрелки (вниз к центру)
    c.line(arrow_top_x, arrow_top_y, arrow_top_x, arrow_top_y - arrow_size)
    # Голова стрелки (треугольник, указывает вниз)
    c.line(arrow_top_x, arrow_top_y - arrow_size, arrow_top_x - arrow_head_size/2, arrow_top_y - arrow_size + arrow_head_size)
    c.line(arrow_top_x, arrow_top_y - arrow_size, arrow_top_x + arrow_head_size/2, arrow_top_y - arrow_size + arrow_head_size)
    
    # Стрелочка внизу (центр по X) - указывает вверх к центру (↑)
    # Рисуем стрелку вверх (к центру)
    c.line(arrow_bottom_x, arrow_bottom_y, arrow_bottom_x, arrow_bottom_y + arrow_size)
    # Голова стрелки (треугольник, указывает вверх)
    c.line(arrow_bottom_x, arrow_bottom_y + arrow_size, arrow_bottom_x - arrow_head_size/2, arrow_bottom_y + arrow_size - arrow_head_size)
    c.line(arrow_bottom_x, arrow_bottom_y + arrow_size, arrow_bottom_x + arrow_head_size/2, arrow_bottom_y + arrow_size - arrow_head_size)
    
    # Стрелочка слева (центр по Y) - указывает вправо к центру (→)
    # Рисуем стрелку вправо (к центру)
    c.line(arrow_left_x, arrow_left_y, arrow_left_x + arrow_size, arrow_left_y)
    # Голова стрелки (треугольник, указывает вправо)
    c.line(arrow_left_x + arrow_size, arrow_left_y, arrow_left_x + arrow_size - arrow_head_size, arrow_left_y - arrow_head_size/2)
    c.line(arrow_left_x + arrow_size, arrow_left_y, arrow_left_x + arrow_size - arrow_head_size, arrow_left_y + arrow_head_size/2)
    
    # Стрелочка справа (центр по Y) - указывает влево к центру (←)
    # Рисуем стрелку влево (к центру)
    c.line(arrow_right_x, arrow_right_y, arrow_right_x - arrow_size, arrow_right_y)
    # Голова стрелки (треугольник, указывает влево)
    c.line(arrow_right_x - arrow_size, arrow_right_y, arrow_right_x - arrow_size + arrow_head_size, arrow_right_y - arrow_head_size/2)
    c.line(arrow_right_x - arrow_size, arrow_right_y, arrow_right_x - arrow_size + arrow_head_size, arrow_right_y + arrow_head_size/2)
    
    # Возвращаем черный цвет для дальнейшей отрисовки
    c.setStrokeColorRGB(0, 0, 0)
    
    # Добавляем копирайт
    add_copyright_to_pdf(c, brand=brand)
    
    c.save()


def calculate_optimal_blocks_per_page(num_blocks, blocks_per_page_min, min_fill_ratio=0.33):
    """
    Рассчитывает оптимальное количество блоков на странице, чтобы страницы были заполнены более чем на min_fill_ratio
    
    Args:
        num_blocks: Общее количество блоков
        blocks_per_page_min: Минимальное количество блоков на странице (например, 56 или 85)
        min_fill_ratio: Минимальный коэффициент заполнения страницы (по умолчанию 0.33 = 1/3)
    
    Returns:
        Оптимальное количество блоков на странице
    """
    # Если блоков меньше или равно минимуму, используем все блоки на одной странице
    if num_blocks <= blocks_per_page_min:
        return num_blocks
    
    # Вычисляем количество страниц с минимальным количеством блоков
    pages = (num_blocks + blocks_per_page_min - 1) // blocks_per_page_min
    
    # Вычисляем количество блоков на последней странице
    blocks_on_last_page = num_blocks % blocks_per_page_min
    if blocks_on_last_page == 0:
        blocks_on_last_page = blocks_per_page_min
    
    # Если последняя страница заполнена менее чем на min_fill_ratio, увеличиваем количество блоков на странице
    if blocks_on_last_page < blocks_per_page_min * min_fill_ratio and pages > 1:
        # Пробуем разные варианты количества страниц, начиная с меньшего количества
        for target_pages in range(pages - 1, 0, -1):
            optimal_blocks = (num_blocks + target_pages - 1) // target_pages
            blocks_on_last = num_blocks % optimal_blocks
            if blocks_on_last == 0:
                blocks_on_last = optimal_blocks
            
            # Проверяем, что последняя страница заполнена более чем на min_fill_ratio
            if blocks_on_last >= optimal_blocks * min_fill_ratio:
                # Проверяем, что не меньше минимума
                if optimal_blocks >= blocks_per_page_min:
                    return optimal_blocks
        
        # Если не нашли оптимальное значение, возвращаем минимум
        return blocks_per_page_min
    
    return blocks_per_page_min


def create_a4_pages(image_path, base_output_name, num_blocks_width=None, num_blocks_height=None, 
                     include_colors=True, include_symbols=True, blocks_per_page_width=None, blocks_per_page_height=None, output_dir=None, palette=None, palette_image_path=None, project_name=None, qr_url=None, painted_cells=None, vertical_lines=None, horizontal_lines=None, auto_blocks_per_page=False, use_dmc=False, brand='marfa'):
    """
    Создает PDF файлы формата A4, разбивая большую таблицу на страницы
    Каждая страница содержит blocks_per_page_width x blocks_per_page_height блоков
    
    Args:
        painted_cells: dict - словарь {(col, row): color} с закрашенными ячейками
        vertical_lines: list - вертикальные линии сетки для определения индексов ячеек
        horizontal_lines: list - горизонтальные линии сетки для определения индексов ячеек
        auto_blocks_per_page: bool - автоматически рассчитывать оптимальное количество блоков на странице
    """
    if num_blocks_width is None:
        num_blocks_width = config.NUM_BLOCKS_WIDTH
    if num_blocks_height is None:
        num_blocks_height = config.NUM_BLOCKS_HEIGHT
    if blocks_per_page_width is None:
        blocks_per_page_width = config.BLOCKS_PER_PAGE_WIDTH
    if blocks_per_page_height is None:
        blocks_per_page_height = config.BLOCKS_PER_PAGE_HEIGHT
    
    # Автоматический расчет оптимального количества блоков на странице
    if auto_blocks_per_page:
        original_width = blocks_per_page_width
        original_height = blocks_per_page_height
        blocks_per_page_width = calculate_optimal_blocks_per_page(num_blocks_width, blocks_per_page_width, min_fill_ratio=0.33)
        blocks_per_page_height = calculate_optimal_blocks_per_page(num_blocks_height, blocks_per_page_height, min_fill_ratio=0.33)
        if blocks_per_page_width != original_width or blocks_per_page_height != original_height:
            print(f"[A4] Автоматический расчет: блоков на странице {blocks_per_page_width}x{blocks_per_page_height} (было {original_width}x{original_height})")
    """
    Создает PDF файлы формата A4, разбивая большую таблицу на страницы
    Каждая страница содержит blocks_per_page_width x blocks_per_page_height блоков
    """
    print(f"[A4] Создание страниц A4 из {image_path}...")
    
    # Загружаем изображение
    img = Image.open(image_path)
    img = img.convert("RGB")
    img_width, img_height = img.size
    
    # Вычисляем размер клетки на основе размера изображения
    cell_width = img_width / num_blocks_width
    cell_height = img_height / num_blocks_height
    
    # Используем переданную палитру или извлекаем из изображения
    if palette is None:
        palette = extract_dominant_colors(image_path, 25)
    char_list = SPECIAL_CHARS * ((len(palette) // len(SPECIAL_CHARS)) + 1)
    color_to_char = {color: char_list[i] for i, color in enumerate(palette)}
    
    # Функция для определения индексов ячейки по координатам пикселя
    def get_cell_indices(img_x, img_y):
        """Определяет индексы ячейки (col, row) по координатам пикселя"""
        if not vertical_lines or not horizontal_lines:
            return None, None
        col = -1
        row = -1
        for i in range(len(vertical_lines) - 1):
            if vertical_lines[i] <= img_x < vertical_lines[i + 1]:
                col = i
                break
        for i in range(len(horizontal_lines) - 1):
            if horizontal_lines[i] <= img_y < horizontal_lines[i + 1]:
                row = i
                break
        if col >= 0 and row >= 0:
            return (col, row)
        return None, None
    
    # Вычисляем количество страниц
    pages_x = (num_blocks_width + blocks_per_page_width - 1) // blocks_per_page_width
    pages_y = (num_blocks_height + blocks_per_page_height - 1) // blocks_per_page_height
    total_pages = pages_x * pages_y
    
    # Используем переданную папку или создаем папку A4-pdf по умолчанию
    if output_dir is None:
        output_dir = "A4-pdf"
    os.makedirs(output_dir, exist_ok=True)
    
    # Извлекаем имя файла из base_output_name
    base_filename = os.path.basename(base_output_name)
    
    # Вычисляем фиксированный размер клетки на основе полного количества блоков на странице
    # Это гарантирует одинаковый размер клеток на всех страницах
    page_width, page_height = A4
    margin = 10 * mm
    available_width = page_width - 2 * margin
    available_height = page_height - 2 * margin
    # Используем полное количество блоков на странице для расчета размера клетки
    cell_size_pt = min(available_width / blocks_per_page_width, available_height / blocks_per_page_height)
    
    page_num = 1
    
    for page_y_idx in range(pages_y):
        for page_x_idx in range(pages_x):
            # Вычисляем диапазон блоков для этой страницы
            start_x_block = page_x_idx * blocks_per_page_width
            end_x_block = min(start_x_block + blocks_per_page_width, num_blocks_width)
            start_y_block = page_y_idx * blocks_per_page_height
            end_y_block = min(start_y_block + blocks_per_page_height, num_blocks_height)
            
            blocks_on_page_x = end_x_block - start_x_block
            blocks_on_page_y = end_y_block - start_y_block
            
            # Создаем PDF страницу формата A4
            output_path = os.path.join(output_dir, f"{base_filename}_a4_page_{page_num}.pdf")
            c = canvas.Canvas(output_path, pagesize=A4)
            set_pdf_metadata(c, f"Схема для вышивания - страница {page_num}")
            
            # Вычисляем размер таблицы на странице
            table_width = cell_size_pt * blocks_on_page_x
            table_height = cell_size_pt * blocks_on_page_y
            
            # Выравниваем таблицу по левому краю (с отступом margin)
            start_x = margin
            # По вертикали выравниваем сверху (с отступом margin)
            start_y = page_height - margin - table_height
            
            # Рисуем блоки для этой страницы
            for local_y in range(blocks_on_page_y):
                global_y = start_y_block + local_y
                for local_x in range(blocks_on_page_x):
                    global_x = start_x_block + local_x
                    
                    # Получаем цвет из изображения
                    pixel_x = int(global_x * cell_width + cell_width / 2)
                    pixel_y = int(global_y * cell_height + cell_height / 2)
                    
                    # Ограничиваем координаты
                    pixel_x = min(pixel_x, img_width - 1)
                    pixel_y = min(pixel_y, img_height - 1)
                    
                    color = img.getpixel((pixel_x, pixel_y))
                    
                    # Определяем индексы ячейки для проверки, закрашена ли она
                    col, row = get_cell_indices(pixel_x, pixel_y)
                    is_painted = (col, row) in painted_cells if (painted_cells and col is not None and row is not None) else True
                    
                    # Находим ближайший цвет из палитры (используем для заливки и символа)
                    closest_color = min(palette, key=lambda c: color_distance(color, c))
                    
                    # Получаем символ только если ячейка закрашена (есть в painted_cells)
                    # Если ячейка стерта (нет в painted_cells), символ не присваиваем
                    if is_painted:
                        symbol = color_to_char.get(closest_color, '')
                    else:
                        symbol = ''  # Стертая ячейка - символ не нужен
                    
                    # Вычисляем позицию клетки в PDF
                    pdf_x = start_x + local_x * cell_size_pt
                    pdf_y = start_y + (blocks_on_page_y - 1 - local_y) * cell_size_pt  # Инвертируем Y
                    
                    # Рисуем цветной блок: закрашенные ячейки — цвет из палитры, без символов — белый фон
                    if include_colors:
                        LIGHT_BEIGE_A4 = (255, 255, 255)  # белый фон для ячеек без символов на A4
                        fill_color = closest_color if is_painted else LIGHT_BEIGE_A4
                        c.setFillColorRGB(fill_color[0]/255.0, fill_color[1]/255.0, fill_color[2]/255.0)
                        c.setStrokeColorRGB(0, 0, 0)
                        c.setLineWidth(0.1)
                        c.rect(pdf_x, pdf_y, cell_size_pt, cell_size_pt, fill=1, stroke=1)
                    
                    # Рисуем символ, если нужно
                    if include_symbols and symbol and cell_size_pt > 1.5 * mm:
                        try:
                            font_size = max(config.MIN_SYMBOL_SIZE_PT, cell_size_pt * config.SYMBOL_SIZE_RATIO)
                            text_x = pdf_x + cell_size_pt / 2
                            text_y = pdf_y + cell_size_pt / 2 - font_size / 3
                            
                            # Определяем контрастный цвет символа в зависимости от яркости фона ячейки
                            if include_colors:
                                # Используем цвет ячейки для определения контрастного цвета текста
                                text_color = get_optimal_text_color(closest_color)
                            else:
                                # Если цвета нет, используем черный по умолчанию
                                text_color = (0, 0, 0)
                            
                            # Устанавливаем цвет текста (значения от 0.0 до 1.0 для reportlab)
                            c.setFillColorRGB(text_color[0]/255.0, text_color[1]/255.0, text_color[2]/255.0)
                            # Используем только базовый шрифт Helvetica
                            c.setFont("Helvetica", font_size)
                            c.drawCentredString(text_x, text_y, symbol)
                        except:
                            pass
            
            # Рисуем тонкую сетку
            c.setStrokeColorRGB(0, 0, 0)
            c.setLineWidth(0.1)
            
            # Вертикальные линии
            for x in range(blocks_on_page_x + 1):
                line_x = start_x + x * cell_size_pt
                c.line(line_x, start_y, line_x, start_y + table_height)
            
            # Горизонтальные линии
            for y in range(blocks_on_page_y + 1):
                line_y = start_y + (blocks_on_page_y - y) * cell_size_pt
                c.line(start_x, line_y, start_x + table_width, line_y)
            
            # Рисуем жирные линии каждые 10 блоков (немного уменьшенная жирность)
            c.setLineWidth(1.5)
            c.setStrokeColorRGB(0, 0, 0)
            
            # Вертикальные жирные линии (каждые 10 блоков)
            for x in range(blocks_on_page_x + 1):
                global_x = start_x_block + x
                if global_x % 10 == 0:
                    line_x = start_x + x * cell_size_pt
                    c.line(line_x, start_y, line_x, start_y + table_height)
            
            # Горизонтальные жирные линии (каждые 10 блоков)
            for y in range(blocks_on_page_y + 1):
                global_y = start_y_block + y
                if global_y % 10 == 0:
                    line_y = start_y + (blocks_on_page_y - y) * cell_size_pt
                    c.line(start_x, line_y, start_x + table_width, line_y)
            
            # Добавляем номера блоков по краям (шкалы по 10) - ПЕРЕД стрелками
            c.setFillColorRGB(0, 0, 0)
            # Размер шрифта адаптируем к размеру клетки и увеличиваем в 1.5 раза
            base_font_size = max(5, min(8, int(cell_size_pt * 0.3)))
            scale_font_name, scale_font_size = get_scale_font(base_font_size * 1.5)
            c.setFont(scale_font_name, scale_font_size)
            
            # Номера по оси X (сверху) - каждые 10 блоков
            for x in range(0, blocks_on_page_x + 1):
                global_x = start_x_block + x
                if global_x % 10 == 0:
                    number_x = start_x + x * cell_size_pt
                    number_text = str(global_x)
                    # Отступ сверху от таблицы
                    number_y = start_y + table_height + 2 * mm
                    c.drawCentredString(number_x, number_y, number_text)
            
            # Номера по оси Y (слева) - каждые 10 блоков
            for y in range(0, blocks_on_page_y + 1):
                global_y = start_y_block + y
                if global_y % 10 == 0 and global_y != 0:  # Пропускаем 0, чтобы не дублировать с осью X
                    line_y = start_y + (blocks_on_page_y - y) * cell_size_pt
                    number_text = str(global_y)
                    # Отступ слева от таблицы
                    number_x = start_x - 3 * mm
                    c.drawRightString(number_x, line_y - scale_font_size / 3, number_text)
            
            # Рисуем красные линии центра (если центр таблицы попадает на эту страницу)
            center_table_x = num_blocks_width // 2  # Центр всей таблицы (75 для 150)
            center_table_y = num_blocks_height // 2  # Центр всей таблицы (75 для 150)
            
            # Проверяем, попадает ли центр таблицы на эту страницу
            if start_x_block <= center_table_x < end_x_block and start_y_block <= center_table_y < end_y_block:
                # Вычисляем локальные координаты центра на странице
                local_center_x = center_table_x - start_x_block
                local_center_y = center_table_y - start_y_block
                
                # Устанавливаем параметры для красных линий
                c.setStrokeColorRGB(1, 0, 0)  # Красный цвет
                c.setLineWidth(1.2)  # Тонкая красная линия для A4
                
                # Вертикальная красная линия через центр
                center_line_x = start_x + local_center_x * cell_size_pt
                c.line(center_line_x, start_y, center_line_x, start_y + table_height)
                
                # Горизонтальная красная линия через центр
                center_line_y = start_y + (blocks_on_page_y - local_center_y) * cell_size_pt
                c.line(start_x, center_line_y, start_x + table_width, center_line_y)
                
                # Рисуем стрелочки на границах страницы - ПОСЛЕ номеров
                arrow_size = 6 * mm
                arrow_line_width = 2.0
                arrow_offset = 8 * mm
                arrow_head_size = 3 * mm
                
                # Вычисляем позиции стрелок
                arrow_top_x = center_line_x
                arrow_top_y = start_y + table_height + arrow_offset
                arrow_bottom_x = center_line_x
                arrow_bottom_y = start_y - arrow_offset
                arrow_left_x = start_x - arrow_offset
                arrow_left_y = center_line_y
                arrow_right_x = start_x + table_width + arrow_offset
                arrow_right_y = center_line_y
                
                # Рисуем белые прямоугольники под стрелками для перекрытия номеров
                c.setFillColorRGB(1, 1, 1)  # Белый цвет
                c.setStrokeColorRGB(1, 1, 1)
                
                # Увеличиваем размеры белого прямоугольника
                white_box_size_horizontal = 12 * mm  # Ширина для горизонтальной стрелки
                white_box_size_vertical = 7 * mm  # Высота для вертикальной стрелки
                white_box_padding = 2 * mm  # Запас вокруг
                
                # Вычисляем размер текста для самого большого номера на странице
                # Используем уже определенные scale_font_name и scale_font_size из номеров выше
                max_global_x = min(start_x_block + blocks_on_page_x, num_blocks_width)
                max_global_y = min(start_y_block + blocks_on_page_y, num_blocks_height)
                max_number = max(max_global_x, max_global_y)
                max_number_text = str(max_number)
                max_text_width = c.stringWidth(max_number_text, scale_font_name, scale_font_size)
                text_height = scale_font_size * 1.3
                
                # Используем максимум из фиксированного размера и размера текста
                white_box_width = max(white_box_size_horizontal, max_text_width + 4 * mm)
                white_box_height = max(white_box_size_vertical, text_height + 2 * mm)
                
                # Белый прямоугольник под стрелкой сверху (перекрывает номер, если он попадает на центр)
                # Номера рисуются на start_y + table_height + 2 * mm
                number_offset_y_a4 = 2 * mm
                c.rect(arrow_top_x - white_box_width/2, 
                       start_y + table_height + number_offset_y_a4 - white_box_padding,
                       white_box_width, 
                       white_box_height, 
                       fill=1, stroke=0)
                
                # Белый прямоугольник под стрелкой слева (перекрывает номер, если он попадает на центр)
                # Номера рисуются на start_x - 3 * mm
                number_offset_x_a4 = 3 * mm
                c.rect(start_x - number_offset_x_a4 - white_box_height - 1 * mm,
                       arrow_left_y - white_box_width/2,
                       white_box_height + 1 * mm,
                       white_box_width,
                       fill=1, stroke=0)
                
                # Теперь рисуем стрелки поверх белых прямоугольников
                c.setStrokeColorRGB(0, 0, 0)
                c.setFillColorRGB(0, 0, 0)
                c.setLineWidth(arrow_line_width)
                
                # Стрелочка вверху (центр по X) - указывает вниз к центру (↓)
                c.line(arrow_top_x, arrow_top_y, arrow_top_x, arrow_top_y - arrow_size)
                c.line(arrow_top_x, arrow_top_y - arrow_size, arrow_top_x - arrow_head_size/2, arrow_top_y - arrow_size + arrow_head_size)
                c.line(arrow_top_x, arrow_top_y - arrow_size, arrow_top_x + arrow_head_size/2, arrow_top_y - arrow_size + arrow_head_size)
                
                # Стрелочка внизу (центр по X) - указывает вверх к центру (↑)
                c.line(arrow_bottom_x, arrow_bottom_y, arrow_bottom_x, arrow_bottom_y + arrow_size)
                c.line(arrow_bottom_x, arrow_bottom_y + arrow_size, arrow_bottom_x - arrow_head_size/2, arrow_bottom_y + arrow_size - arrow_head_size)
                c.line(arrow_bottom_x, arrow_bottom_y + arrow_size, arrow_bottom_x + arrow_head_size/2, arrow_bottom_y + arrow_size - arrow_head_size)
                
                # Стрелочка слева (центр по Y) - указывает вправо к центру (→)
                c.line(arrow_left_x, arrow_left_y, arrow_left_x + arrow_size, arrow_left_y)
                c.line(arrow_left_x + arrow_size, arrow_left_y, arrow_left_x + arrow_size - arrow_head_size, arrow_left_y - arrow_head_size/2)
                c.line(arrow_left_x + arrow_size, arrow_left_y, arrow_left_x + arrow_size - arrow_head_size, arrow_left_y + arrow_head_size/2)
                
                # Стрелочка справа (центр по Y) - указывает влево к центру (←)
                c.line(arrow_right_x, arrow_right_y, arrow_right_x - arrow_size, arrow_right_y)
                c.line(arrow_right_x - arrow_size, arrow_right_y, arrow_right_x - arrow_size + arrow_head_size, arrow_right_y - arrow_head_size/2)
                c.line(arrow_right_x - arrow_size, arrow_right_y, arrow_right_x - arrow_size + arrow_head_size, arrow_right_y + arrow_head_size/2)
                
                # Возвращаем черный цвет для дальнейшей отрисовки
                c.setStrokeColorRGB(0, 0, 0)
            
            # Добавляем копирайт (при DMC — бренд Lilu&Stitch; иначе по brand)
            add_copyright_to_pdf(c, use_dmc=use_dmc, brand=brand)
            
            c.save()
            page_num += 1
    
    # Создаем отдельную страницу с палитрой (как последнюю страницу)
    if palette_image_path and os.path.exists(palette_image_path):
        from reportlab.lib.utils import ImageReader
        
        palette_page_path = os.path.join(output_dir, f"{base_filename}_a4_page_{total_pages + 1}.pdf")
        c = canvas.Canvas(palette_page_path, pagesize=A4)
        set_pdf_metadata(c, "Палитра цветов")
        
        page_width, page_height = A4
        # В fcrossy используется 300 DPI для изображений и PIL (координаты сверху вниз)
        # В reportlab координаты снизу вверх, поэтому нужно пересчитывать
        
        # Параметры как в fcrossy (в мм, потом конвертируем)
        # В fcrossy: MARGINS["a4_top_text"] = 1 мм, MARGINS["a4_side"] = 5 мм, MARGINS["a4_between_images"] = 1.5 мм
        dpi = 300
        margin_top_text_mm = 1  # Отступ сверху в мм (как в fcrossy)
        margin_side_mm = 5  # Отступ сбоку в мм (как в fcrossy)
        margin_between_mm = 1.5  # Отступ между элементами в мм (как в fcrossy)
        
        # Конвертируем мм в точки для reportlab (1 мм = 72/25.4 точек)
        margin_top_text = margin_top_text_mm * mm
        margin_side = margin_side_mm * mm
        margin_between = margin_between_mm * mm
        
        # В reportlab координаты идут снизу вверх, поэтому начинаем сверху страницы
        current_y = page_height - margin_top_text
        dpi_ratio = 72.0 / 300.0  # Конвертация из пикселей (300 DPI) в точки
        
        # --- sor2.png (если есть) ---
        # Пробуем разные пути к sor2.png
        try:
            from utils.path_utils import get_static_path
            sor_image_path = get_static_path("sor2.png")
            if not os.path.exists(sor_image_path):
                sor_image_path = get_static_path("pic/sor2.png")
        except ImportError:
            sor_image_path = os.path.join("static", "sor2.png")
            if not os.path.exists(sor_image_path):
                sor_image_path = os.path.join("static", "pic", "sor2.png")
        if os.path.exists(sor_image_path):
            try:
                sor_img = Image.open(sor_image_path)
                sor_width_px, sor_height_px = sor_img.size
                
                # Масштабируем по ширине страницы
                # width_px = int(a4_width_inch * dpi) = int(8.27 * 300) = 2481 пикселей
                a4_width_px = int(8.27 * dpi)  # Ширина A4 в пикселях при 300 DPI
                margin_side_px = int(margin_side_mm * dpi / 25.4)  # margin_side в пикселях
                sor_ratio = (a4_width_px - 2 * margin_side_px) / sor_width_px
                sor_width = sor_width_px * dpi_ratio * sor_ratio
                sor_height = sor_height_px * dpi_ratio * sor_ratio
                sor_x = (page_width - sor_width) / 2
                sor_y = current_y - sor_height  # В reportlab вычитаем высоту (верх изображения в current_y)
                
                # Вставляем изображение
                c.drawImage(ImageReader(sor_img), sor_x, sor_y, width=sor_width, height=sor_height)
                current_y = sor_y - margin_between  # Двигаемся вверх в координатах (вниз по странице)
            except Exception as e:
                print(f"[ERROR] Не удалось вставить sor2.png: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"[WARNING] sor2.png не найден ни в одном из путей:")
            print(f"[WARNING]   - static/sor2.png")
            print(f"[WARNING]   - static/pic/sor2.png")
        
        # Бар с текстом "Ключ к набору" убран по запросу пользователя
        
        # --- Информационный текст: при DMC — на английском ---
        info_text = "Cross in two threads" if use_dmc else "Крест в две нитки"
        font_size_info = config.INFO_FONT_SIZE
        c.setFillColorRGB(0, 0, 0)  # Черный цвет
        # Используем шрифт из static/fonts
        info_font_name, info_font_size = get_info_font(font_size_info)
        c.setFont(info_font_name, info_font_size)
        
        info_text_x = margin_side
        # В reportlab drawString рисует текст снизу вверх, поэтому вычитаем только размер шрифта
        info_text_y = current_y - info_font_size
        c.drawString(info_text_x, info_text_y, info_text)
        
        current_y = info_text_y - margin_between  # Двигаемся вверх в координатах (вниз по странице)
        
        # --- Вставка палитры (легенды) ---
        palette_img = Image.open(palette_image_path)
        img_width, img_height = palette_img.size
        
        # Область для легенды - выравнивание по левому краю
        # Используем тот же отступ слева, что и для остального контента (margin_side)
        dpi_ratio = 72.0 / 300.0
        # Позиционируем по левому краю с отступом margin_side
        legend_area_left = margin_side
        # Ширина области - доступная ширина страницы минус отступы
        available_width = page_width - 2 * margin_side
        # Высота области должна быть ограничена доступным пространством
        # Вычисляем доступную высоту от current_y до низа страницы
        available_height_from_current = current_y - (10 * mm)  # Минимальный отступ снизу 10mm
        available_height = min(available_height_from_current, 500 * dpi_ratio)
        
        # Вычисляем коэффициент масштабирования
        # img_width и img_height в пикселях при 300 DPI, конвертируем в точки
        img_width_pt = img_width * dpi_ratio
        img_height_pt = img_height * dpi_ratio
        ratio = min(available_width / img_width_pt, available_height / img_height_pt)
        new_width = img_width_pt * ratio
        new_height = img_height_pt * ratio
        
        # Позиционируем по левому краю
        x = legend_area_left
        y = current_y - new_height  # В reportlab вычитаем высоту
        
        # Проверяем, что палитра не выходит за пределы страницы
        min_y = 10 * mm  # Минимальный отступ снизу
        if y < min_y:
            print(f"[WARNING] Палитра выходит за пределы страницы (y={y/mm:.1f}mm), корректируем позицию")
            y = min_y
            # Если палитра не помещается, уменьшаем её размер
            if new_height > (current_y - min_y):
                new_height = current_y - min_y
                new_width = img_width_pt * (new_height / img_height_pt)
        
        # Вставляем изображение палитры
        c.drawImage(ImageReader(palette_img), x, y, width=new_width, height=new_height)
        
        # При DMC на странице легенды QR-коды не вставляются
        if not use_dmc:
            # Добавляем QR-коды под легендой:
            # 1) "Электронная схема" – только если указан qr_url
            # 2) "Поддержка" (Telegram) – всегда
            # 3) "Промокоды на вышивку" (Telegram) – всегда
            # 4) "Поддержка / промокоды (MAX)" – всегда
            try:
                from export.qr_code_generator import generate_qr_without_text
                import tempfile

                # Общие параметры размещения
                qr_margin = 15 * mm  # Отступ между легендой и QR-кодами
                text_height_above_qr = 5 * mm  # Место для текста над QR-кодом
                qr_gap_mm = 20  # Расстояние между первыми тремя QR-кодами
                qr_gap_before_4th_mm = 28  # Больший отступ перед 4-м QR, чтобы подпись не налезала на подпись 3-го

                # Базовая позиция по X/Y (левый край – по X)
                base_qr_x = x
                if base_qr_x < margin_side:
                    base_qr_x = margin_side

                # Эти переменные будем заполнять по мере добавления QR
                row_y = None          # общая высота ряда QR-кодов
                next_qr_x = None      # X для следующего QR

                # --- 1. Первый QR: "Электронная схема" (только если есть qr_url) ---
                if qr_url and qr_url.strip():
                    try:
                        temp_qr_path = os.path.join(tempfile.gettempdir(), "qr_palette_temp.jpg")
                        qr_path = generate_qr_without_text(
                            url=qr_url,
                            output_path=temp_qr_path,
                            qr_box_size=2,  # Маленькие ячейки QR-кода
                            qr_border=2,
                            qr_version=5  # Маленькая версия для компактного размера
                        )
                        if qr_path and os.path.exists(qr_path):
                            qr_img = Image.open(qr_path).convert("RGB")
                            qr_width_pt = qr_img.width
                            qr_height_pt = qr_img.height

                            # Вычисляем Y-координату для ряда QR-кодов
                            row_y = y - qr_height_pt - qr_margin - text_height_above_qr
                            if row_y < 10 * mm:
                                row_y = 10 * mm

                            qr_x = base_qr_x

                            # Вставляем первый QR-код
                            c.drawImage(ImageReader(qr_img), qr_x, row_y, width=qr_width_pt, height=qr_height_pt)

                            # Текст над первым QR
                            text_qr1 = "Электронная схема"
                            qr_text_font_name, qr_text_font_size = get_scale_font(9)
                            c.setFont(qr_text_font_name, qr_text_font_size)
                            c.setFillColorRGB(0, 0, 0)
                            text_width_qr1 = c.stringWidth(text_qr1, qr_text_font_name, qr_text_font_size)
                            text_x_qr1 = qr_x + (qr_width_pt - text_width_qr1) / 2
                            text_y_qr1 = row_y + qr_height_pt + 2 * mm
                            c.drawString(text_x_qr1, text_y_qr1, text_qr1)

                            print(f"[INFO] QR-код инструкции добавлен (размер: {qr_width_pt/mm:.1f}mm x {qr_height_pt/mm:.1f}mm)")
                            try:
                                os.remove(qr_path)
                            except:
                                pass

                            # X для следующего QR (справа от первого)
                            next_qr_x = qr_x + qr_width_pt + qr_gap_mm * mm
                    except Exception as e_first_qr:
                        print(f"[WARNING] Ошибка при добавлении первого QR-кода (электронная схема): {e_first_qr}")
                        import traceback
                        traceback.print_exc()

                # --- 2. Второй QR: "Поддержка" (Telegram) – всегда ---
                try:
                    temp_tg_qr_path = os.path.join(tempfile.gettempdir(), "qr_telegram_temp.jpg")
                    tg_qr_path = generate_qr_without_text(
                        url="https://t.me/+gOLwQEcrbZcyNWVi",
                        output_path=temp_tg_qr_path,
                        qr_box_size=2,
                        qr_border=2,
                        qr_version=5
                    )
                    if tg_qr_path and os.path.exists(tg_qr_path):
                        tg_qr_img = Image.open(tg_qr_path).convert("RGB")
                        tg_qr_width_pt = tg_qr_img.width
                        tg_qr_height_pt = tg_qr_img.height

                        # Если первый QR не отрисован – задаем row_y и next_qr_x по второму
                        if row_y is None:
                            row_y = y - tg_qr_height_pt - qr_margin - text_height_above_qr
                            if row_y < 10 * mm:
                                row_y = 10 * mm
                        if next_qr_x is None:
                            next_qr_x = base_qr_x

                        tg_qr_x = next_qr_x
                        tg_qr_y = row_y

                        # Проверяем правую границу страницы
                        if tg_qr_x + tg_qr_width_pt > page_width - margin_side:
                            tg_qr_x = max(margin_side, page_width - margin_side - tg_qr_width_pt)

                        # Рисуем второй QR
                        c.drawImage(ImageReader(tg_qr_img), tg_qr_x, tg_qr_y, width=tg_qr_width_pt, height=tg_qr_height_pt)

                        # Текст над вторым QR
                        text_qr2 = "Поддержка (TG)"
                        qr_text_font_name_2, qr_text_font_size_2 = get_scale_font(9)
                        c.setFont(qr_text_font_name_2, qr_text_font_size_2)
                        c.setFillColorRGB(0, 0, 0)
                        text_width_qr2 = c.stringWidth(text_qr2, qr_text_font_name_2, qr_text_font_size_2)
                        text_x_qr2 = tg_qr_x + (tg_qr_width_pt - text_width_qr2) / 2
                        text_y_qr2 = tg_qr_y + tg_qr_height_pt + 2 * mm
                        c.drawString(text_x_qr2, text_y_qr2, text_qr2)

                        print(f"[INFO] QR-код Поддержка добавлен (размер: {tg_qr_width_pt/mm:.1f}mm x {tg_qr_height_pt/mm:.1f}mm)")
                        try:
                            os.remove(tg_qr_path)
                        except:
                            pass

                        # X для третьего QR
                        next_qr_x = tg_qr_x + tg_qr_width_pt + qr_gap_mm * mm
                except Exception as e_tg:
                    print(f"[WARNING] Ошибка при добавлении QR-кода Telegram (Поддержка): {e_tg}")
                    import traceback
                    traceback.print_exc()

                # --- 3. Третий QR: "Группа с акциями и новинками" (Telegram) – всегда ---
                try:
                    temp_tg_qr3_path = os.path.join(tempfile.gettempdir(), "qr_telegram_promo_temp.jpg")
                    tg_qr3_path = generate_qr_without_text(
                        url="https://t.me/+gOLwQEcrbZcyNWVi",
                        output_path=temp_tg_qr3_path,
                        qr_box_size=2,
                        qr_border=2,
                        qr_version=5
                    )
                    if tg_qr3_path and os.path.exists(tg_qr3_path):
                        tg_qr3_img = Image.open(tg_qr3_path).convert("RGB")
                        tg_qr3_width_pt = tg_qr3_img.width
                        tg_qr3_height_pt = tg_qr3_img.height

                        # Если до этого ни одного QR не отрисовано, задаем row_y и X
                        if row_y is None:
                            row_y = y - tg_qr3_height_pt - qr_margin - text_height_above_qr
                            if row_y < 10 * mm:
                                row_y = 10 * mm
                        if next_qr_x is None:
                            next_qr_x = base_qr_x

                        tg_qr3_x = next_qr_x
                        tg_qr3_y = row_y

                        # Проверяем правую границу страницы
                        if tg_qr3_x + tg_qr3_width_pt > page_width - margin_side:
                            tg_qr3_x = max(margin_side, page_width - margin_side - tg_qr3_width_pt)

                        # Вставляем третий QR-код
                        c.drawImage(ImageReader(tg_qr3_img), tg_qr3_x, tg_qr3_y, width=tg_qr3_width_pt, height=tg_qr3_height_pt)

                        # Текст над третьим QR
                        text_qr3 = "Промокоды на вышивку (TG)"
                        qr_text_font_name_3, qr_text_font_size_3 = get_scale_font(9)
                        c.setFont(qr_text_font_name_3, qr_text_font_size_3)
                        c.setFillColorRGB(0, 0, 0)
                        text_width_qr3 = c.stringWidth(text_qr3, qr_text_font_name_3, qr_text_font_size_3)
                        text_x_qr3 = tg_qr3_x + (tg_qr3_width_pt - text_width_qr3) / 2
                        text_y_qr3 = tg_qr3_y + tg_qr3_height_pt + 2 * mm
                        c.drawString(text_x_qr3, text_y_qr3, text_qr3)

                        print(f"[INFO] QR-код 'Группа с акциями и новинками' добавлен (размер: {tg_qr3_width_pt/mm:.1f}mm x {tg_qr3_height_pt/mm:.1f}mm)")
                        try:
                            os.remove(tg_qr3_path)
                        except:
                            pass
                        # X для четвёртого QR (увеличенный отступ, чтобы подпись не налезала на 3-й)
                        next_qr_x = tg_qr3_x + tg_qr3_width_pt + qr_gap_before_4th_mm * mm
                except Exception as e_tg3:
                    print(f"[WARNING] Ошибка при добавлении третьего QR-кода Telegram (акции и новинки): {e_tg3}")
                    import traceback
                    traceback.print_exc()

                # --- 4. Четвёртый QR: "Поддержка / промокоды (MAX)" – всегда ---
                try:
                    temp_max_qr_path = os.path.join(tempfile.gettempdir(), "qr_max_temp.jpg")
                    max_qr_path = generate_qr_without_text(
                        url="https://max.ru/join/uhXMWZ8AiUo0aPlhL8jRZ76_-uUSJapTbtbKteen2Cs",
                        output_path=temp_max_qr_path,
                        qr_box_size=2,
                        qr_border=2,
                        qr_version=5
                    )
                    if max_qr_path and os.path.exists(max_qr_path):
                        max_qr_img = Image.open(max_qr_path).convert("RGB")
                        max_qr_width_pt = max_qr_img.width
                        max_qr_height_pt = max_qr_img.height

                        if row_y is None:
                            row_y = y - max_qr_height_pt - qr_margin - text_height_above_qr
                            if row_y < 10 * mm:
                                row_y = 10 * mm
                        if next_qr_x is None:
                            next_qr_x = base_qr_x

                        max_qr_x = next_qr_x
                        max_qr_y = row_y

                        if max_qr_x + max_qr_width_pt > page_width - margin_side:
                            max_qr_x = max(margin_side, page_width - margin_side - max_qr_width_pt)

                        c.drawImage(ImageReader(max_qr_img), max_qr_x, max_qr_y, width=max_qr_width_pt, height=max_qr_height_pt)

                        text_qr4 = "Поддержка / промокоды (MAX)"
                        qr_text_font_name_4, qr_text_font_size_4 = get_scale_font(9)
                        c.setFont(qr_text_font_name_4, qr_text_font_size_4)
                        c.setFillColorRGB(0, 0, 0)
                        text_width_qr4 = c.stringWidth(text_qr4, qr_text_font_name_4, qr_text_font_size_4)
                        text_x_qr4 = max_qr_x + (max_qr_width_pt - text_width_qr4) / 2
                        text_y_qr4 = max_qr_y + max_qr_height_pt + 2 * mm
                        c.drawString(text_x_qr4, text_y_qr4, text_qr4)

                        print(f"[INFO] QR-код 'Поддержка / промокоды (MAX)' добавлен (размер: {max_qr_width_pt/mm:.1f}mm x {max_qr_height_pt/mm:.1f}mm)")
                        try:
                            os.remove(max_qr_path)
                        except:
                            pass
                except Exception as e_max:
                    print(f"[WARNING] Ошибка при добавлении четвёртого QR-кода (MAX): {e_max}")
                    import traceback
                    traceback.print_exc()

            except Exception as e:
                print(f"[WARNING] Ошибка при добавлении QR-кодов под легендой: {e}")
                import traceback
                traceback.print_exc()
        
        # Добавляем копирайт (при DMC — бренд Lilu&Stitch; иначе по brand)
        add_copyright_to_pdf(c, use_dmc=use_dmc, brand=brand)
        
        c.save()
        total_pages += 1
    
    return total_pages


def add_palette_page_to_pdf(pdf_path, palette_image_path, palette, color_to_char, brand='marfa'):
    """
    Добавляет страницу с палитрой в существующий PDF файл
    Вставляет sor2.png, bar с текстом "Ключ к набору" и палитру
    """
    if PdfMerger is None:
        print(f"[WARNING] PyPDF2/pypdf не установлен. Пропускаем добавление палитры в PDF.")
        return False
    
    try:
        from reportlab.lib.utils import ImageReader
        
        # Создаем временный PDF с палитрой
        temp_palette_pdf = pdf_path.replace('.pdf', '_palette_temp.pdf')
        c = canvas.Canvas(temp_palette_pdf, pagesize=A4)
        set_pdf_metadata(c, "Палитра цветов")
        page_width, page_height = A4
        
        # Параметры как в create_a4_pages
        dpi = 300
        margin_top_text_mm = 1
        margin_side_mm = 5
        margin_between_mm = 1.5
        
        margin_top_text = margin_top_text_mm * mm
        margin_side = margin_side_mm * mm
        margin_between = margin_between_mm * mm
        dpi_ratio = 72.0 / 300.0
        
        current_y = page_height - margin_top_text
        
        # --- sor2.png (если есть) ---
        try:
            from utils.path_utils import get_static_path
            sor_image_path = get_static_path("sor2.png")
            if not os.path.exists(sor_image_path):
                sor_image_path = get_static_path("pic/sor2.png")
        except ImportError:
            sor_image_path = os.path.join("static", "sor2.png")
            if not os.path.exists(sor_image_path):
                sor_image_path = os.path.join("static", "pic", "sor2.png")
        if os.path.exists(sor_image_path):
            try:
                sor_img = Image.open(sor_image_path)
                sor_width_px, sor_height_px = sor_img.size
                a4_width_px = int(8.27 * dpi)
                margin_side_px = int(margin_side_mm * dpi / 25.4)
                sor_ratio = (a4_width_px - 2 * margin_side_px) / sor_width_px
                sor_width = sor_width_px * dpi_ratio * sor_ratio
                sor_height = sor_height_px * dpi_ratio * sor_ratio
                sor_x = (page_width - sor_width) / 2
                sor_y = current_y - sor_height
                c.drawImage(ImageReader(sor_img), sor_x, sor_y, width=sor_width, height=sor_height)
                current_y = sor_y - margin_between
            except Exception as e:
                print(f"[WARNING] Не удалось вставить sor2.png: {e}")
        
        # Бар с текстом "Ключ к набору" убран по запросу пользователя
        
        # --- Информационный текст "Крест в две нитки" ---
        info_text = "Крест в две нитки"
        font_size_info = config.INFO_FONT_SIZE
        c.setFillColorRGB(0, 0, 0)
        # Используем шрифт из static/fonts
        info_font_name, info_font_size = get_info_font(font_size_info)
        c.setFont(info_font_name, info_font_size)
        info_text_x = margin_side
        info_text_y = current_y - info_font_size
        c.drawString(info_text_x, info_text_y, info_text)
        current_y = info_text_y - margin_between
        
        # --- Вставка палитры (легенды) ---
        if os.path.exists(palette_image_path):
            try:
                palette_img = Image.open(palette_image_path)
                img_width, img_height = palette_img.size
                
                legend_area_left_px = 60
                legend_area_width_px = 1500 - 60
                legend_area_height_px = 500
                
                legend_area_left = legend_area_left_px * dpi_ratio
                available_width = legend_area_width_px * dpi_ratio
                available_height = legend_area_height_px * dpi_ratio
                
                img_width_pt = img_width * dpi_ratio
                img_height_pt = img_height * dpi_ratio
                ratio = min(available_width / img_width_pt, available_height / img_height_pt)
                new_width = img_width_pt * ratio
                new_height = img_height_pt * ratio
                
                x = legend_area_left
                y = current_y - new_height
                
                # Проверяем, что палитра не выходит за пределы страницы
                if y < 0:
                    print(f"[WARNING] Палитра выходит за пределы страницы (y={y/mm:.1f}mm), корректируем позицию")
                    y = 10 * mm  # Минимальный отступ снизу
                
                c.drawImage(ImageReader(palette_img), x, y, width=new_width, height=new_height)
            except Exception as e:
                print(f"[WARNING] Не удалось вставить изображение палитры: {e}")
        else:
            print(f"[WARNING] Изображение палитры не найдено: {palette_image_path}")
        
        # Добавляем копирайт
        add_copyright_to_pdf(c, brand=brand)
        
        c.save()
        
        # Объединяем оригинальный PDF с палитрой
        merger = PdfMerger()
        set_pdf_merger_metadata(merger, "Схема для вышивания с палитрой")
        
        # Добавляем оригинальный PDF
        if os.path.exists(pdf_path):
            merger.append(pdf_path)
        
        # Добавляем страницу с палитрой
        merger.append(temp_palette_pdf)
        
        # Сохраняем объединенный PDF
        merger.write(pdf_path)
        merger.close()
        
        # Удаляем временный файл
        if os.path.exists(temp_palette_pdf):
            os.remove(temp_palette_pdf)
        
        return True
        
    except Exception as e:
        print(f"[ERROR] Ошибка при добавлении палитры в PDF: {e}")
        import traceback
        traceback.print_exc()
        return False


def add_table_page_to_pdf(pdf_path, excel_path, use_dmc=False, colorize_symbols=True, brand='marfa'):
    """
    Добавляет страницу с таблицей из Excel файла в существующий PDF файл
    Таблица содержит колонки: N, Symbol, Gamma/DMC, Length
    
    Args:
        use_dmc: если True, использует столбец DMC вместо Gamma и название колонки "DMC"
        colorize_symbols: если True, закрашивает ячейки Symbols цветом Gamma (по умолчанию True)
    """
    if PdfMerger is None:
        print(f"[WARNING] PyPDF2/pypdf не установлен. Пропускаем добавление таблицы в PDF.")
        return False
    
    if pd is None:
        print(f"[WARNING] pandas не установлен. Пропускаем добавление таблицы в PDF.")
        return False
    
    try:
        if not os.path.exists(excel_path):
            print(f"[WARNING] Excel файл не найден: {excel_path}")
            return False
        
        # Создаем временный PDF с таблицей
        temp_table_pdf = pdf_path.replace('.pdf', '_table_temp.pdf')
        c = canvas.Canvas(temp_table_pdf, pagesize=A4)
        set_pdf_metadata(c, "Таблица соответствия цветов")
        page_width, page_height = A4
        
        # Параметры
        dpi = 300
        margin_top_text_mm = 1
        margin_side_mm = 5
        margin_between_mm = 1.5
        
        margin_top_text = margin_top_text_mm * mm
        margin_side = margin_side_mm * mm
        margin_between = margin_between_mm * mm
        dpi_ratio = 72.0 / 300.0
        
        current_y = page_height - margin_top_text
        
        # --- sor2.png (если есть) ---
        try:
            from utils.path_utils import get_static_path
            sor_image_path = get_static_path("sor2.png")
            if not os.path.exists(sor_image_path):
                sor_image_path = get_static_path("pic/sor2.png")
        except ImportError:
            sor_image_path = os.path.join("static", "sor2.png")
            if not os.path.exists(sor_image_path):
                sor_image_path = os.path.join("static", "pic", "sor2.png")
        if os.path.exists(sor_image_path):
            try:
                from reportlab.lib.utils import ImageReader
                sor_img = Image.open(sor_image_path)
                sor_width_px, sor_height_px = sor_img.size
                a4_width_px = int(8.27 * dpi)
                margin_side_px = int(margin_side_mm * dpi / 25.4)
                sor_ratio = (a4_width_px - 2 * margin_side_px) / sor_width_px
                sor_width = sor_width_px * dpi_ratio * sor_ratio
                sor_height = sor_height_px * dpi_ratio * sor_ratio
                sor_x = (page_width - sor_width) / 2
                sor_y = current_y - sor_height
                c.drawImage(ImageReader(sor_img), sor_x, sor_y, width=sor_width, height=sor_height)
                current_y = sor_y - margin_between
            except Exception as e:
                print(f"[WARNING] Не удалось вставить sor2.png: {e}")
        
        # Бар с текстом "Ключ к набору" убран по запросу пользователя
        
        # --- Информационный текст: при DMC — на английском ---
        info_text = "Cross in two threads" if use_dmc else "Крест в две нитки"
        font_size_info = config.INFO_FONT_SIZE
        c.setFillColorRGB(0, 0, 0)
        info_font_name, info_font_size = get_info_font(font_size_info)
        c.setFont(info_font_name, info_font_size)
        info_text_x = margin_side
        info_text_y = current_y - info_font_size
        c.drawString(info_text_x, info_text_y, info_text)
        current_y = info_text_y - margin_between * 2
        
        # --- Создание таблицы из Excel ---
        print(f"[TABLE] Загрузка данных из Excel для таблицы...")
        df = pd.read_excel(excel_path, engine='openpyxl')
        
        # Сортируем по столбцу N, если он есть
        if 'N' in df.columns:
            df = df.sort_values('N')
        
        # Определяем название колонки и столбца для использования
        column_name = "DMC" if use_dmc else "Gamma"
        column_key = "DMC" if use_dmc else "Gamma"
        
        # Параметры таблицы
        table_start_x = margin_side
        table_start_y = current_y
        row_height = 8 * mm
        col_width_n = 15 * mm
        col_width_symbol = 20 * mm
        col_width_gamma = 30 * mm
        col_width_length = 20 * mm
        
        # Заголовки таблицы
        header_font_size = config.TABLE_HEADER_FONT_SIZE
        c.setFont("Helvetica-Bold", header_font_size)
        c.setFillColorRGB(0.9, 0.9, 0.9)  # Светло-серый фон для заголовков
        
        # Рисуем заголовки
        c.rect(table_start_x, table_start_y - row_height, col_width_n, row_height, fill=1, stroke=1)
        c.rect(table_start_x + col_width_n, table_start_y - row_height, col_width_symbol, row_height, fill=1, stroke=1)
        c.rect(table_start_x + col_width_n + col_width_symbol, table_start_y - row_height, col_width_gamma, row_height, fill=1, stroke=1)
        c.rect(table_start_x + col_width_n + col_width_symbol + col_width_gamma, table_start_y - row_height, col_width_length, row_height, fill=1, stroke=1)
        
        c.setFillColorRGB(0, 0, 0)
        # Выравниваем заголовки по центру
        header_y = table_start_y - row_height + row_height / 2 - header_font_size / 3
        c.drawCentredString(table_start_x + col_width_n / 2, header_y, "N")
        c.drawCentredString(table_start_x + col_width_n + col_width_symbol / 2, header_y, "Symbol")
        c.drawCentredString(table_start_x + col_width_n + col_width_symbol + col_width_gamma / 2, header_y, column_name)
        c.drawCentredString(table_start_x + col_width_n + col_width_symbol + col_width_gamma + col_width_length / 2, header_y, "Length")
        
        # Данные таблицы
        data_font_size = config.TABLE_DATA_FONT_SIZE
        c.setFont("Helvetica", data_font_size)
        current_table_y = table_start_y - row_height
        
        for idx, row in df.iterrows():
            current_table_y -= row_height
            
            # Проверяем, не вышли ли за пределы страницы
            if current_table_y < 30 * mm:
                # Добавляем копирайт на текущую страницу перед переходом на новую
                add_copyright_to_pdf(c, use_dmc=use_dmc, brand=brand)
                # Начинаем новую страницу
                c.showPage()
                current_table_y = page_height - margin_top_text - row_height
                
                # Рисуем заголовки на новой странице
                c.setFont("Helvetica-Bold", header_font_size)
                c.setFillColorRGB(0.9, 0.9, 0.9)
                c.rect(table_start_x, current_table_y, col_width_n, row_height, fill=1, stroke=1)
                c.rect(table_start_x + col_width_n, current_table_y, col_width_symbol, row_height, fill=1, stroke=1)
                c.rect(table_start_x + col_width_n + col_width_symbol, current_table_y, col_width_gamma, row_height, fill=1, stroke=1)
                c.rect(table_start_x + col_width_n + col_width_symbol + col_width_gamma, current_table_y, col_width_length, row_height, fill=1, stroke=1)
                c.setFillColorRGB(0, 0, 0)
                # Выравниваем заголовки по центру на новой странице
                header_y = current_table_y + row_height / 2 - header_font_size / 3
                c.drawCentredString(table_start_x + col_width_n / 2, header_y, "N")
                c.drawCentredString(table_start_x + col_width_n + col_width_symbol / 2, header_y, "Symbol")
                c.drawCentredString(table_start_x + col_width_n + col_width_symbol + col_width_gamma / 2, header_y, column_name)
                c.drawCentredString(table_start_x + col_width_n + col_width_symbol + col_width_gamma + col_width_length / 2, header_y, "Length")
                c.setFont("Helvetica", data_font_size)
                current_table_y -= row_height
            
            # Получаем данные
            n_val = str(int(row['N'])) if 'N' in df.columns and pd.notna(row['N']) else str(idx + 1)
            symbol_val = str(row.get('Символ', '')).strip() if pd.notna(row.get('Символ', '')) else ''
            # Используем DMC или Gamma в зависимости от use_dmc
            if use_dmc and 'DMC' in df.columns:
                number_val = str(row.get('DMC', '')).strip() if pd.notna(row.get('DMC', '')) else ''
            else:
                number_val = str(row.get('Gamma', '')).strip() if pd.notna(row.get('Gamma', '')) else ''
            length_val = str(row.get('Длина', '')).strip() if pd.notna(row.get('Длина', '')) else ''
            
            # Получаем RGB цвет Gamma для закрашивания ячейки Symbols
            gamma_rgb = None
            if 'Цвет Гамма' in df.columns and pd.notna(row.get('Цвет Гамма', '')):
                # Пытаемся получить из hex
                gamma_hex = str(row.get('Цвет Гамма', '')).strip()
                if gamma_hex and gamma_hex.startswith('#'):
                    try:
                        hex_color = gamma_hex.lstrip('#')
                        if len(hex_color) >= 6:
                            r = int(hex_color[0:2], 16)
                            g = int(hex_color[2:4], 16)
                            b = int(hex_color[4:6], 16)
                            gamma_rgb = (r, g, b)
                    except:
                        pass
            
            # Если не получили из hex, пробуем из R, G, B столбцов
            if gamma_rgb is None and 'R' in df.columns and 'G' in df.columns and 'B' in df.columns:
                try:
                    r_val = row.get('R', 0)
                    g_val = row.get('G', 0)
                    b_val = row.get('B', 0)
                    if pd.notna(r_val) and pd.notna(g_val) and pd.notna(b_val):
                        r = int(r_val)
                        g = int(g_val)
                        b = int(b_val)
                        gamma_rgb = (r, g, b)
                except:
                    pass
            
            # Рисуем ячейки
            c.rect(table_start_x, current_table_y, col_width_n, row_height, fill=0, stroke=1)
            
            # Ячейка Symbols - используем логику из create_pdf_table (только если colorize_symbols=True)
            if colorize_symbols and gamma_rgb and symbol_val:
                # Закрашиваем ячейку цветом Gamma (как в create_pdf_table)
                r, g, b = gamma_rgb
                c.setFillColorRGB(r/255.0, g/255.0, b/255.0)
                c.setStrokeColorRGB(0, 0, 0)  # Черная рамка
                c.setLineWidth(0.1)  # Тонкая линия
                c.rect(table_start_x + col_width_n, current_table_y, col_width_symbol, row_height, fill=1, stroke=1)
                
                # Используем get_optimal_text_color для выбора цвета текста (как в create_pdf_table)
                text_color = get_optimal_text_color(gamma_rgb)
                c.setFillColorRGB(text_color[0]/255.0, text_color[1]/255.0, text_color[2]/255.0)
                
                # Рисуем символ по центру ячейки (как в create_pdf_table)
                font_size = max(8, row_height * 0.7)  # Адаптируем размер шрифта под высоту строки
                text_x = table_start_x + col_width_n + col_width_symbol / 2
                text_y = current_table_y + row_height / 2 - font_size / 3
                c.setFont("Helvetica", font_size)
                c.drawCentredString(text_x, text_y, symbol_val)
                
                # Возвращаем стандартную толщину линии для остальных ячеек
                c.setLineWidth(1.0)
            else:
                # Если не нужно закрашивать или нет цвета Gamma или символа, рисуем обычную ячейку
                c.rect(table_start_x + col_width_n, current_table_y, col_width_symbol, row_height, fill=0, stroke=1)
                c.setFillColorRGB(0, 0, 0)  # Черный текст
                if symbol_val:
                    text_y = current_table_y + row_height / 2 - data_font_size / 3
                    c.drawCentredString(table_start_x + col_width_n + col_width_symbol / 2, text_y, symbol_val)
            
            c.rect(table_start_x + col_width_n + col_width_symbol, current_table_y, col_width_gamma, row_height, fill=0, stroke=1)
            c.rect(table_start_x + col_width_n + col_width_symbol + col_width_gamma, current_table_y, col_width_length, row_height, fill=0, stroke=1)
            
            # Рисуем текст в остальных ячейках (выравниваем по центру)
            c.setFillColorRGB(0, 0, 0)  # Возвращаем черный цвет для остальных ячеек
            text_y = current_table_y + row_height / 2 - data_font_size / 3
            c.drawCentredString(table_start_x + col_width_n / 2, text_y, n_val)
            # Символ уже нарисован выше (если был цвет Gamma)
            if not (colorize_symbols and gamma_rgb and symbol_val):
                c.drawCentredString(table_start_x + col_width_n + col_width_symbol / 2, text_y, symbol_val)
            c.drawCentredString(table_start_x + col_width_n + col_width_symbol + col_width_gamma / 2, text_y, number_val)
            c.drawCentredString(table_start_x + col_width_n + col_width_symbol + col_width_gamma + col_width_length / 2, text_y, length_val)
        
        # Добавляем копирайт (при DMC — бренд Lilu&Stitch; иначе по brand)
        add_copyright_to_pdf(c, use_dmc=use_dmc, brand=brand)
        
        c.save()
        
        # Объединяем оригинальный PDF с таблицей
        merger = PdfMerger()
        set_pdf_merger_metadata(merger, "Схема для вышивания с таблицей")
        
        # Добавляем оригинальный PDF
        if os.path.exists(pdf_path):
            merger.append(pdf_path)
        
        # Добавляем страницу с таблицей
        merger.append(temp_table_pdf)
        
        # Сохраняем объединенный PDF
        merger.write(pdf_path)
        merger.close()
        
        # Удаляем временный файл
        if os.path.exists(temp_table_pdf):
            os.remove(temp_table_pdf)
        
        print(f"[OK] Таблица добавлена в PDF: {pdf_path}")
        return True
        
    except Exception as e:
        print(f"[ERROR] Ошибка при добавлении таблицы в PDF: {e}")
        import traceback
        traceback.print_exc()
        return False


def find_closest_gamma_color(palette_color, gamma_excel_path):
    """
    Находит ближайший цвет Gamma для цвета из палитры
    Сравнивает с цветами из столбца "Цвет Гамма" (или R, G, B) в Excel файле
    Возвращает номер Gamma и RGB цвет
    """
    if pd is None:
        return None, None, float('inf')
    
    try:
        # Загружаем Excel файл
        gamma_df = pd.read_excel(gamma_excel_path)
        
        # Ищем столбец с цветом Gamma (может быть в разных вариантах названия)
        color_column = None
        # Проверяем различные варианты названий столбца
        possible_names = ['Цвет Гамма', 'Цвет_Гамма', 'Цвет', 'Цвет Гамма (XSPro)', 'Цвет_github']
        for name in possible_names:
            if name in gamma_df.columns:
                color_column = name
                break
        
        # Если нет столбца с цветом, используем R, G, B
        use_rgb = color_column is None
        
        min_distance = float('inf')
        closest_gamma = None
        closest_rgb = None
        
        for _, row in gamma_df.iterrows():
            if use_rgb:
                # Используем R, G, B столбцы
                if 'R' in row and 'G' in row and 'B' in row:
                    r = int(row['R']) if pd.notna(row['R']) else 0
                    g = int(row['G']) if pd.notna(row['G']) else 0
                    b = int(row['B']) if pd.notna(row['B']) else 0
                    gamma_rgb = (r, g, b)
                else:
                    continue
            else:
                # Используем столбец с hex цветом
                color_hex = str(row[color_column]).strip()
                # Проверяем, что это не NaN
                if pd.isna(row[color_column]) or color_hex == 'nan':
                    continue
                if color_hex.startswith('#'):
                    try:
                        hex_color = color_hex.lstrip('#')
                        if len(hex_color) >= 6:
                            r = int(hex_color[0:2], 16)
                            g = int(hex_color[2:4], 16)
                            b = int(hex_color[4:6], 16)
                            gamma_rgb = (r, g, b)
                        else:
                            continue
                    except:
                        continue
                else:
                    continue
            
            # Вычисляем расстояние между цветами
            distance = color_distance(palette_color, gamma_rgb)
            
            if distance < min_distance:
                min_distance = distance
                closest_gamma = row.get('Gamma', '')
                closest_rgb = gamma_rgb
        
        return closest_gamma, closest_rgb, min_distance
        
    except Exception as e:
        print(f"[WARNING] Ошибка при поиске ближайшего цвета Gamma: {e}")
        return None, None, float('inf')


def find_dmc_by_gamma_number(gamma_number, gamma_excel_path):
    """
    Находит номер DMC по номеру Gamma напрямую из Excel файла
    Ищет в столбце "Gamma" и возвращает соответствующее значение из столбца "DMC"
    
    Args:
        gamma_number: номер Gamma (может быть строкой или числом)
        gamma_excel_path: путь к Excel файлу с данными Gamma/DMC
    
    Returns:
        str или None: номер DMC или None, если не найден
    """
    if pd is None:
        return None
    
    try:
        # Загружаем Excel файл
        gamma_df = pd.read_excel(gamma_excel_path)
        
        # Нормализуем номер Gamma для сравнения
        normalized_gamma = normalize_gamma(gamma_number)
        if not normalized_gamma:
            return None
        
        # Ищем столбец "Gamma"
        if 'Gamma' not in gamma_df.columns:
            print(f"[WARNING] Столбец 'Gamma' не найден в Excel файле")
            return None
        
        # Ищем столбец "DMC"
        if 'DMC' not in gamma_df.columns:
            print(f"[WARNING] Столбец 'DMC' не найден в Excel файле")
            return None
        
        # Ищем строку с совпадающим номером Gamma
        for _, row in gamma_df.iterrows():
            row_gamma = row.get('Gamma', '')
            if pd.notna(row_gamma):
                # Нормализуем номер Gamma из строки для сравнения
                normalized_row_gamma = normalize_gamma(row_gamma)
                if normalized_row_gamma == normalized_gamma:
                    # Нашли совпадение, возвращаем DMC
                    dmc_value = row.get('DMC', '')
                    if pd.notna(dmc_value):
                        # Преобразуем в строку и очищаем
                        dmc_str = str(dmc_value).strip()
                        # Удаляем .0 в конце, если это число
                        if dmc_str.endswith('.0'):
                            dmc_str = dmc_str[:-2]
                        return dmc_str if dmc_str else None
        
        return None
        
    except Exception as e:
        print(f"[WARNING] Ошибка при поиске DMC по номеру Gamma: {e}")
        return None


def save_gamma_legend_to_excel(palette, color_to_char, gamma_excel_path, output_excel_path, color_counts=None, use_dmc=False):
    """
    Сохраняет данные палитры в Excel файл <артикул>.xlsx
    Колонки: N (нумерация), Цвет (RGB), Цвет Гамма (hex), Gamma, DMC (если use_dmc=True), Символ, Длина
    
    Args:
        use_dmc: если True, добавляет столбец DMC, найденный по номеру Gamma
    """
    if pd is None:
        print("[ERROR] pandas не установлен. Невозможно сохранить в Excel.")
        return False
    
    try:
        
        # Подготавливаем данные для сохранения
        data = []
        for idx, color in enumerate(palette):
            # Нормализуем цвет до кортежа
            if isinstance(color, (list, np.ndarray)):
                normalized_color = tuple(int(c) for c in color)
            elif isinstance(color, tuple):
                normalized_color = tuple(int(c) for c in color)
            else:
                normalized_color = color
            
            # Получаем символ
            symbol = color_to_char.get(normalized_color, '')
            
            # Находим ближайший цвет Gamma
            gamma_num = None
            gamma_rgb = None
            gamma_hex = None
            if gamma_excel_path and os.path.exists(gamma_excel_path):
                gamma_num, gamma_rgb, distance = find_closest_gamma_color(normalized_color, gamma_excel_path)
                if gamma_rgb:
                    gamma_hex = f"#{gamma_rgb[0]:02x}{gamma_rgb[1]:02x}{gamma_rgb[2]:02x}"
            
            # Находим DMC по номеру Gamma, если use_dmc=True
            dmc_num = None
            if use_dmc and gamma_num and gamma_excel_path and os.path.exists(gamma_excel_path):
                dmc_num = find_dmc_by_gamma_number(gamma_num, gamma_excel_path)
            
            # Форматируем цвет RGB как строку
            color_rgb_str = f"{normalized_color[0]},{normalized_color[1]},{normalized_color[2]}"
            
            # Рассчитываем длину нитки (как в органайзере)
            length_value = ""
            if color_counts and normalized_color in color_counts:
                blocks_count = color_counts[normalized_color]
                # Длина = количество кубиков * 0.004 (метры)
                import math
                length_meters = blocks_count * 0.004
                # Если < 0.5 - округление вверх без +1, если >= 0.5 - округление вверх +1
                if length_meters < 0.5:
                    length_rounded = math.ceil(length_meters)
                else:
                    length_rounded = math.ceil(length_meters) + 1
                length_value = str(int(length_rounded))
            
            row_data = {
                'N': idx + 1,  # Нумерация начиная с 1
                'Цвет': color_rgb_str,
                'R': normalized_color[0],
                'G': normalized_color[1],
                'B': normalized_color[2],
                'Цвет Гамма': gamma_hex if gamma_hex else '',
                'Gamma': gamma_num if gamma_num else '',
                'Символ': symbol,
                'Длина': length_value
            }
            
            # Добавляем столбец DMC, если use_dmc=True
            if use_dmc:
                row_data['DMC'] = dmc_num if dmc_num else ''
            
            data.append(row_data)
        
        # Создаем DataFrame
        df = pd.DataFrame(data)
        
        # Упорядочиваем колонки: N первым
        column_order = ['N', 'Цвет', 'R', 'G', 'B', 'Цвет Гамма', 'Gamma']
        if use_dmc:
            column_order.append('DMC')
        column_order.extend(['Символ', 'Длина'])
        # Оставляем только те колонки, которые есть в данных
        column_order = [col for col in column_order if col in df.columns]
        df = df[column_order]
        
        # Сохраняем в Excel
        df.to_excel(output_excel_path, index=False, engine='openpyxl')
        
        # Закрашиваем ячейки в колонке "Цвет Гамма" соответствующими цветами
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            
            # Загружаем созданный файл
            wb = load_workbook(output_excel_path)
            ws = wb.active
            
            # Находим индекс колонки "Цвет Гамма"
            header_row = 1
            gamma_color_col_idx = None
            for col_idx, cell in enumerate(ws[header_row], start=1):
                if cell.value == 'Цвет Гамма':
                    gamma_color_col_idx = col_idx
                    break
            
            # Если колонка найдена, закрашиваем ячейки
            if gamma_color_col_idx:
                colored_count = 0
                for row_idx in range(2, len(df) + 2):  # Начинаем с 2-й строки (после заголовка)
                    cell = ws.cell(row=row_idx, column=gamma_color_col_idx)
                    hex_value = cell.value
                    
                    # Проверяем, что значение не пустое и имеет формат hex
                    if hex_value and isinstance(hex_value, str) and hex_value.startswith('#'):
                        try:
                            # Убираем # и конвертируем hex в RGB
                            hex_color = hex_value.lstrip('#').upper()
                            if len(hex_color) >= 6:
                                # В openpyxl PatternFill использует формат RRGGBB без #, но может быть проблема с длиной
                                # Убеждаемся, что берем ровно 6 символов
                                hex_color_clean = hex_color[:6]
                                
                                # Проверяем, что это валидный hex
                                int(hex_color_clean, 16)
                                
                                # Создаем заливку с цветом (openpyxl использует формат RRGGBB без #)
                                fill = PatternFill(start_color=hex_color_clean, 
                                                  end_color=hex_color_clean, 
                                                  fill_type="solid")
                                cell.fill = fill
                                colored_count += 1
                        except (ValueError, IndexError, TypeError) as e:
                            # Если не удалось распарсить hex, пропускаем
                            pass
                
                if colored_count > 0:
                    print(f"[EXCEL] Закрашено {colored_count} ячеек в колонке 'Цвет Гамма'")
            else:
                print(f"[WARNING] Колонка 'Цвет Гамма' не найдена в Excel файле")
            
            # Сохраняем изменения
            wb.save(output_excel_path)
            wb.close()
        except Exception as e:
            print(f"[WARNING] Не удалось закрасить ячейки в Excel: {e}")
            # Продолжаем выполнение, даже если закрашивание не удалось
        
        return True
        
    except Exception as e:
        print(f"[ERROR] Ошибка при сохранении в Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


def load_gamma_legend_from_excel(excel_path):
    """
    Загружает данные из Excel файла <артикул>.xlsx
    Сортирует по столбцу N (нумерация) и возвращает словарь color_to_char, список палитры в порядке нумерации и словарь dmc_numbers
    
    Returns:
        tuple: (color_to_char, palette, numbers, dmc_numbers) или (None, None, None, None) при ошибке
    """
    if pd is None:
        print("[ERROR] pandas не установлен. Невозможно загрузить из Excel.")
        return None, None, None, None
    
    try:
        if not os.path.exists(excel_path):
            print(f"[WARNING] Файл {excel_path} не найден")
            return None, None, None, None
        
        print(f"[GAMMA-LEGEND] Загрузка данных из {excel_path}...")
        
        # Загружаем Excel файл с явным указанием обработки Unicode
        df = pd.read_excel(excel_path, engine='openpyxl', dtype={'Символ': str})
        
        # Сортируем по столбцу N (нумерация), если он есть
        if 'N' in df.columns:
            df = df.sort_values('N')
        else:
            print(f"[WARNING] Столбец N не найден, используем порядок из файла")
        
        # Создаем словарь color_to_char
        color_to_char = {}
        palette = []
        numbers = []  # Список номеров N из Excel
        dmc_numbers = {}  # Словарь цвет -> DMC номер
        
        for _, row in df.iterrows():
            # Получаем номер N из Excel
            n_number = int(row['N']) if 'N' in df.columns and pd.notna(row['N']) else len(palette) + 1
            numbers.append(n_number)
            
            # Получаем цвет RGB
            if 'R' in df.columns and 'G' in df.columns and 'B' in df.columns:
                r = int(row['R']) if pd.notna(row['R']) else 0
                g = int(row['G']) if pd.notna(row['G']) else 0
                b = int(row['B']) if pd.notna(row['B']) else 0
                color = (r, g, b)
            elif 'Цвет' in df.columns:
                # Парсим строку формата "R,G,B"
                color_str = str(row['Цвет']).strip()
                try:
                    r, g, b = map(int, color_str.split(','))
                    color = (r, g, b)
                except:
                    continue
            else:
                continue
            
            # Получаем символ из столбца "Символ" Excel файла
            symbol_raw = row.get('Символ', '')
            if pd.notna(symbol_raw):
                # Преобразуем в строку и очищаем от пробелов
                symbol = str(symbol_raw).strip()
                
                # Убеждаемся, что символ правильно декодирован (Unicode)
                # Если символ выглядит как байтовая строка, декодируем
                if isinstance(symbol, bytes):
                    try:
                        symbol = symbol.decode('utf-8')
                    except:
                        try:
                            symbol = symbol.decode('latin-1')
                        except:
                            symbol = str(symbol_raw)
                
                # Проверяем, что символ не пустой и не NaN
                if symbol == '' or symbol == 'nan' or symbol.lower() == 'none':
                    symbol = ''
            else:
                symbol = ''
            
            # Проверяем, что символ не пустой
            if not symbol:
                print(f"[WARNING] Пустой символ для цвета {color} (N={n_number}) в Excel файле")
            else:
                # Проверяем кодировку символа
                try:
                    # Пробуем закодировать в UTF-8, чтобы убедиться, что это валидный Unicode
                    symbol.encode('utf-8')
                    symbol_repr = repr(symbol)
                    # Отладочная информация для всех символов (первые 10 подробно, остальные кратко)
                    if len(palette) < 10:
                        print(f"[GAMMA-LEGEND] Загружен символ для цвета {color} (N={n_number}): '{symbol}' (repr: {symbol_repr}, длина: {len(symbol)}, код: {ord(symbol[0]) if symbol else 'N/A'})")
                    elif len(palette) < 30:  # Для остальных выводим кратко
                        print(f"[GAMMA-LEGEND] N={n_number}: '{symbol}' (код: {ord(symbol[0]) if symbol else 'N/A'})")
                except UnicodeEncodeError as e:
                    print(f"[ERROR] Проблема с кодировкой символа для цвета {color} (N={n_number}): {e}")
                    symbol = ''  # Очищаем невалидный символ
            
            # Получаем DMC из столбца DMC, если он есть
            dmc_value = None
            if 'DMC' in df.columns:
                dmc_raw = row.get('DMC', '')
                if pd.notna(dmc_raw):
                    dmc_str = str(dmc_raw).strip()
                    # Удаляем .0 в конце, если это число
                    if dmc_str.endswith('.0'):
                        dmc_str = dmc_str[:-2]
                    if dmc_str and dmc_str != 'nan':
                        dmc_value = dmc_str
            
            # Добавляем в словарь и палитру (в порядке нумерации из Excel)
            # Порядок соответствует столбцу N (нумерация) из Excel
            color_to_char[color] = symbol
            palette.append(color)
            if dmc_value:
                dmc_numbers[color] = dmc_value
        
        return color_to_char, palette, numbers, dmc_numbers
        
    except Exception as e:
        print(f"[ERROR] Ошибка при загрузке из Excel: {e}")
        import traceback
        traceback.print_exc()
        return None, None




def normalize_gamma(value):
    """
    Нормализует значение Gamma для сравнения.
    Создает органайзер с цветами и номерами Gamma
    """
    import pandas as pd
    if pd.isna(value):
        return None
    # Обрабатываем числа: если это float без дробной части, преобразуем в int
    if isinstance(value, (int, float)):
        if value == int(value):
            value = int(value)
    result = str(value).strip().lstrip('0').replace(' ', '').upper()
    # Удаляем .0 в конце, если оно там есть
    if result.endswith('.0'):
        result = result[:-2]
    # Если после удаления ведущих нулей осталась пустая строка или только точка, возвращаем исходное значение как строку
    if not result or result == '.':
        result = str(value).strip().replace(' ', '').upper()
    return result


def calculate_color_counts(layout_image_path, palette, num_blocks_width=150, num_blocks_height=150, painted_cells=None):
    """
    Вычисляет количество кубиков каждого цвета
    Если передан painted_cells, считает напрямую из него (более точно после ручных правок)
    Иначе считает из layout изображения
    """
    from collections import Counter
    import numpy as np
    
    # Подсчитываем количество блоков каждого цвета
    color_counts = {color: 0 for color in palette}
    
    # Если есть painted_cells, используем его для подсчета (более точно)
    if painted_cells and len(painted_cells) > 0:
        print(f"[ORGANIZER] Подсчет количества кубиков из painted_cells...")
        for (col, row), cell_color in painted_cells.items():
            # Нормализуем цвет из painted_cells
            if isinstance(cell_color, (list, np.ndarray)):
                color_tuple = tuple(int(c) for c in cell_color[:3])
            elif isinstance(cell_color, tuple):
                color_tuple = tuple(int(c) for c in cell_color[:3])
            else:
                color_tuple = cell_color
            
            # Находим ближайший цвет из палитры
            min_distance = float('inf')
            closest_color = None
            for color in palette:
                distance = sum((a - b) ** 2 for a, b in zip(color_tuple, color))
                if distance < min_distance:
                    min_distance = distance
                    closest_color = color
            
            # Увеличиваем счетчик для найденного цвета
            if closest_color is not None:
                color_counts[closest_color] += 1
    else:
        # Fallback: считаем из изображения (старый метод)
        print(f"[ORGANIZER] Подсчет количества кубиков из layout изображения...")
        img = Image.open(layout_image_path)
        img = img.convert("RGB")
        img_width, img_height = img.size
        
        # Вычисляем размер одного блока (кубика)
        block_width = img_width / num_blocks_width
        block_height = img_height / num_blocks_height
        
        # Проходим по каждому блоку
        for y_block in range(num_blocks_height):
            for x_block in range(num_blocks_width):
                # Вычисляем координаты центра блока
                center_x = int(x_block * block_width + block_width / 2)
                center_y = int(y_block * block_height + block_height / 2)
                
                # Ограничиваем координаты
                center_x = min(center_x, img_width - 1)
                center_y = min(center_y, img_height - 1)
                
                # Получаем цвет пикселя в центре блока
                block_color = img.getpixel((center_x, center_y))
                
                # Пропускаем белые пиксели (незакрашенные блоки)
                if block_color == (255, 255, 255):
                    continue
                
                # Находим ближайший цвет из палитры
                min_distance = float('inf')
                closest_color = None
                for color in palette:
                    distance = sum((a - b) ** 2 for a, b in zip(block_color, color))
                    if distance < min_distance:
                        min_distance = distance
                        closest_color = color
                
                # Увеличиваем счетчик для найденного цвета
                if closest_color is not None:
                    color_counts[closest_color] += 1
    
    print(f"[ORGANIZER] Подсчитано кубиков для {len(palette)} цветов")
    for color, count in color_counts.items():
        if count > 0:
            print(f"[ORGANIZER]   Цвет RGB{color}: {count} кубиков")
    return color_counts


def create_organizer_image(palette, color_to_char, output_path, project_name="Проект", project_name_text=None, gamma_excel_path=None, layout_image_path=None, num_blocks_width=150, num_blocks_height=150, qr_url=None, painted_cells=None, use_dmc=False):
    """
    Создает органайзер на основе палитры цветов с сопоставлением цветов Gamma или DMC
    Сравнивает цвета палитры с цветами из Excel файла и отображает номер Gamma или DMC
    Использует шаблон organizer_template.png и накладывает organizer_numbers.png
    По образцу fcrossy_3.10
    
    Args:
        use_dmc: если True, использует DMC вместо Gamma для отображения номеров
    """
    print(f"[ORGANIZER] Создание органайзера...")
    if use_dmc:
        print(f"[ORGANIZER] Использование DMC вместо Gamma")
    
    # Если указан путь к Excel файлу, загружаем данные Gamma
    gamma_data = {}
    if gamma_excel_path and os.path.exists(gamma_excel_path):
        print(f"[ORGANIZER] Загрузка данных Gamma из {gamma_excel_path}...")
        for i, color in enumerate(palette):
            gamma_num, gamma_rgb, distance = find_closest_gamma_color(color, gamma_excel_path)
            if gamma_num is not None:
                gamma_normalized = normalize_gamma(gamma_num)
                # Находим DMC по номеру Gamma, если use_dmc=True
                dmc_num = None
                if use_dmc:
                    dmc_num = find_dmc_by_gamma_number(gamma_num, gamma_excel_path)
                
                gamma_data[color] = {
                    'gamma': gamma_normalized if gamma_normalized else str(gamma_num),
                    'dmc': dmc_num if dmc_num else '',
                    'rgb': gamma_rgb,
                    'distance': distance
                }
                if use_dmc and dmc_num:
                    print(f"[ORGANIZER] Цвет {i+1}: RGB{color} -> Gamma {gamma_data[color]['gamma']} -> DMC {dmc_num} (расстояние: {distance:.2f})")
                else:
                    print(f"[ORGANIZER] Цвет {i+1}: RGB{color} -> Gamma {gamma_data[color]['gamma']} (расстояние: {distance:.2f})")
            else:
                gamma_data[color] = {'gamma': '', 'dmc': '', 'rgb': None, 'distance': float('inf')}
    else:
        print(f"[WARNING] Excel файл Gamma не указан или не найден. Органайзер будет создан без номеров Gamma.")
        if gamma_excel_path:
            print(f"[WARNING] Путь: {gamma_excel_path}")
    
    # Вычисляем количество кубиков каждого цвета для расчета длины нитки
    color_counts = {}
    if painted_cells and len(painted_cells) > 0:
        # Используем painted_cells для точного подсчета (учитывает удаленные ячейки)
        color_counts = calculate_color_counts(layout_image_path, palette, num_blocks_width, num_blocks_height, painted_cells=painted_cells)
    elif layout_image_path and os.path.exists(layout_image_path):
        # Fallback: считаем из изображения
        color_counts = calculate_color_counts(layout_image_path, palette, num_blocks_width, num_blocks_height)
    else:
        print(f"[WARNING] Layout изображение не указано и painted_cells отсутствует, длина нитки не будет рассчитана")
    
    # Загружаем шаблон органайзера
    try:
        from utils.path_utils import get_static_path
        organizer_template_path = get_static_path("organizer_template.png")
        if not os.path.exists(organizer_template_path):
            organizer_template_path = get_static_path("pic/organizer_template.png")
    except ImportError:
        organizer_template_path = os.path.join("static", "organizer_template.png")
        if not os.path.exists(organizer_template_path):
            organizer_template_path = os.path.join("static", "pic", "organizer_template.png")
    
    if not os.path.exists(organizer_template_path):
        raise FileNotFoundError(f"Шаблон органайзера не найден: {organizer_template_path}")
    
    print(f"[ORGANIZER] Использование шаблона: {organizer_template_path}")
    img = Image.open(organizer_template_path).convert("RGBA")
    draw = ImageDraw.Draw(img)
    
    # Параметры как в fcrossy
    rect_h = 100  # Высота прямоугольника
    rect_len = 150  # Длина прямоугольника
    shift_before_15 = 0  # Сдвиг до 15-го элемента
    shift_after_15 = 60  # Сдвиг после 15-го элемента
    title_area = (180, 10, 740, 50)  # Область заголовка
    y_coords = [70, 220, 360, 510, 655, 800, 940, 1095, 1245, 1390,
                1545, 1685, 1840, 1985, 2134]  # Y координаты элементов
    x_coords = [235, 455]  # X координаты колонок
    
    # Загружаем шрифты
    font_size = 36
    title_font_size = 42
    font = None
    title_font = None
    
    try:
        from utils.path_utils import get_static_path
        font_path = get_static_path("fonts/Montserrat-SemiBold.ttf")
        if not os.path.exists(font_path):
            font_path = get_static_path("fonts/MontserratAlternates-Regular.otf")
        
        if os.path.exists(font_path):
            font = ImageFont.truetype(font_path, font_size)
            title_font = ImageFont.truetype(font_path, title_font_size)
        else:
            font = get_font(font_size)
            title_font = get_font(title_font_size)
    except ImportError:
        font_path = os.path.join("static", "fonts", "Montserrat-SemiBold.ttf")
        if not os.path.exists(font_path):
            font_path = os.path.join("static", "fonts", "MontserratAlternates-Regular.otf")
        if os.path.exists(font_path):
            font = ImageFont.truetype(font_path, font_size)
            title_font = ImageFont.truetype(font_path, title_font_size)
        else:
            font = get_font(font_size)
            title_font = get_font(title_font_size)
    except Exception:
        font = ImageFont.load_default()
        title_font = ImageFont.load_default()
    
    # Убеждаемся, что шрифты определены
    if font is None:
        font = ImageFont.load_default()
    if title_font is None:
        title_font = ImageFont.load_default()
    
    # Добавляем заголовок (project_name) сверху
    bbox_title = draw.textbbox((0, 0), project_name, font=title_font)
    title_w = bbox_title[2] - bbox_title[0]
    title_h = bbox_title[3] - bbox_title[1]
    title_x = title_area[0] + (title_area[2] - title_area[0] - title_w) // 2
    title_y = title_area[1] + (title_area[3] - title_area[1] - title_h) // 2
    draw.text((title_x, title_y), project_name, font=title_font, fill="black")
    
    # Обрабатываем каждую запись
    row_idx = 0
    for col_idx, x_coord in enumerate(x_coords):
        for y in y_coords:
            if row_idx >= len(palette):
                break
            
            color = palette[row_idx]
            
            # Определяем сдвиг
            shift_x = shift_before_15 if row_idx < 15 else shift_after_15
            
            # Получаем данные Gamma
            gamma_info = gamma_data.get(color, {})
            gamma_num = gamma_info.get('gamma', '')
            gamma_rgb = gamma_info.get('rgb', color)
            
            # Используем DMC вместо Gamma, если use_dmc=True
            if use_dmc:
                display_number = gamma_info.get('dmc', '')
            else:
                display_number = gamma_num
            
            # Вычисляем длину нитки
            length_value = ""
            if color in color_counts:
                blocks_count = color_counts[color]
                # Длина = количество кубиков * 0.004 (метры)
                length_meters = blocks_count * 0.004
                # Если < 0.5 - округление вверх без +1, если >= 0.5 - округление вверх +1
                if length_meters < 0.5:
                    length_rounded = math.ceil(length_meters)
                else:
                    length_rounded = math.ceil(length_meters) + 1
                length_value = str(int(length_rounded))
                print(f"[ORGANIZER] Цвет {row_idx+1}: {blocks_count} кубиков -> {length_value} метров")
            
            # Рисуем цветной прямоугольник
            draw.rectangle(
                [x_coord + shift_x, y, x_coord + rect_len + shift_x, y + rect_h],
                fill=gamma_rgb if gamma_rgb else color,
                outline="black"
            )
            
            # Создаем белый треугольник
            triangle_layer = Image.new("RGBA", img.size, (255, 255, 255, 0))
            triangle_draw = ImageDraw.Draw(triangle_layer)
            
            if row_idx < 15:
                triangle_points = [
                    (x_coord + shift_x, y),
                    (x_coord + shift_x, y + rect_h),
                    (x_coord + rect_h / 2 + shift_x, y + rect_h / 2)
                ]
            else:
                triangle_points = [
                    (x_coord + rect_len + shift_x, y),
                    (x_coord + rect_len + shift_x, y + rect_h),
                    (x_coord + rect_len - rect_h / 2 + shift_x, y + rect_h / 2)
                ]
            
            triangle_draw.polygon(triangle_points, fill=(255, 255, 255, 255))
            img.paste(triangle_layer, (0, 0), triangle_layer)
            
            # Определяем цвет текста на основе яркости фона
            r, g, b = gamma_rgb if gamma_rgb else color
            luminance = 0.299 * r + 0.587 * g + 0.114 * b
            text_color = "white" if luminance < 128 else "black"
            
            # Создаем изображение с текстом
            text_img = Image.new("RGBA", (rect_h, rect_len), (255, 255, 255, 0))
            text_draw = ImageDraw.Draw(text_img)
            
            # Номер GAMMA или DMC
            number_value = display_number if display_number else ""
            bbox_number = text_draw.textbbox((0, 0), number_value, font=font)
            number_w = bbox_number[2] - bbox_number[0]
            number_h = bbox_number[3] - bbox_number[1]
            text_draw.text(((rect_h - number_w) // 2, 5), number_value, font=font, fill=text_color)
            
            # Длина
            bbox_length = text_draw.textbbox((0, 0), length_value, font=font)
            length_w = bbox_length[2] - bbox_length[0]
            length_h = bbox_length[3] - bbox_length[1]
            text_draw.text(((rect_h - length_w) // 2, 5 + number_h + 5), length_value, font=font, fill=text_color)
            
            # Поворачиваем текст
            angle = -90 if col_idx == 0 else 90
            text_img = text_img.rotate(angle, expand=True)
            paste_x = x_coord + shift_x + (rect_len - text_img.width) // 2
            paste_y = y + (rect_h - text_img.height) // 2
            img.paste(text_img, (paste_x, paste_y), text_img)
            
            row_idx += 1
    
    # Накладываем слой с цифрами (organizer_numbers.png)
    try:
        from utils.path_utils import get_static_path
        organizer_numbers_path = get_static_path("organizer_numbers.png")
        if not os.path.exists(organizer_numbers_path):
            organizer_numbers_path = get_static_path("pic/organizer_numbers.png")
    except ImportError:
        organizer_numbers_path = os.path.join("static", "organizer_numbers.png")
        if not os.path.exists(organizer_numbers_path):
            organizer_numbers_path = os.path.join("static", "pic", "organizer_numbers.png")
    
    if os.path.exists(organizer_numbers_path):
        print(f"[ORGANIZER] Наложение цифр: {organizer_numbers_path}")
        try:
            numbers_img = Image.open(organizer_numbers_path).convert("RGBA")
            if numbers_img.size != img.size:
                numbers_img = numbers_img.resize(img.size, Image.NEAREST)
            img.paste(numbers_img, (0, 0), numbers_img)
        except Exception as e:
            print(f"[WARNING] Не удалось наложить цифры: {e}")
    else:
        print(f"[WARNING] Изображение с цифрами не найдено: {organizer_numbers_path}")
    
    # QR-код теперь добавляется только на страницу с легендой (палитрой)
    # Не добавляем QR-код на органайзер
    
    # Добавляем название проекта внизу органайзера (если указано)
    if project_name_text and project_name_text.strip():
        img_width, img_height = img.size
        # Область для текста внизу (примерно 50 пикселей от низа)
        bottom_area = (180, img_height - 60, 740, img_height - 10)
        bottom_text = project_name_text.strip()
        
        # Используем тот же шрифт, что и для заголовка
        bbox_bottom = draw.textbbox((0, 0), bottom_text, font=title_font)
        bottom_w = bbox_bottom[2] - bbox_bottom[0]
        bottom_h = bbox_bottom[3] - bbox_bottom[1]
        bottom_x = bottom_area[0] + (bottom_area[2] - bottom_area[0] - bottom_w) // 2
        bottom_y = bottom_area[1] + (bottom_area[3] - bottom_area[1] - bottom_h) // 2
        draw.text((bottom_x, bottom_y), bottom_text, font=title_font, fill="black")
    
    # Сохраняем в формате JPG с метаданными
    img_rgb = img.convert("RGB")
    save_jpg_with_metadata(img_rgb, output_path, quality=config.JPEG_QUALITY, dpi=(config.DPI, config.DPI), title="Органайзер")


def merge_pdf_pages(output_dir, base_filename, total_pages, merged_output_path):
    """
    Объединяет отдельные страницы PDF в один файл
    """
    if PdfMerger is None:
        print(f"[WARNING] PyPDF2/pypdf не установлен. Пропускаем объединение PDF.")
        return False
    
    print(f"[MERGE] Объединение {total_pages} страниц в один PDF...")
    
    try:
        merger = PdfMerger()
        set_pdf_merger_metadata(merger, f"Схема для вышивания - {base_filename}")
        
        # Добавляем страницы в порядке их создания
        for page_num in range(1, total_pages + 1):
            page_path = os.path.join(output_dir, f"{base_filename}_a4_page_{page_num}.pdf")
            if os.path.exists(page_path):
                merger.append(page_path)
                print(f"[MERGE] Добавлена страница {page_num}/{total_pages}")
            else:
                print(f"[WARNING] Файл не найден: {page_path}")
        
        # Сохраняем объединенный PDF
        merger.write(merged_output_path)
        merger.close()
        
        return True
        
    except Exception as e:
        print(f"[ERROR] Ошибка при объединении PDF: {e}")
        import traceback
        traceback.print_exc()
        return False


def process_single_image(input_file, num_colors=None, num_blocks_width=None, num_blocks_height=None):
    """
    Обрабатывает одно изображение.
    
    :param input_file: путь к входному файлу изображения
    :param num_colors: количество цветов в палитре (если None, берется из config)
    :param num_blocks_width: количество блоков по ширине (если None, берется из config)
    :param num_blocks_height: количество блоков по высоте (если None, берется из config)
    """
    if num_colors is None:
        num_colors = config.NUM_COLORS
    if num_blocks_width is None:
        num_blocks_width = config.NUM_BLOCKS_WIDTH
    if num_blocks_height is None:
        num_blocks_height = config.NUM_BLOCKS_HEIGHT
    """
    Обрабатывает одно изображение.
    
    :param input_file: путь к входному файлу изображения
    :param num_colors: количество цветов в палитре
    :param num_blocks_width: количество блоков по ширине
    :param num_blocks_height: количество блоков по высоте
    """
    # Проверяем существование входного файла
    if not os.path.exists(input_file):
        print(f"[ERROR] Файл {input_file} не найден!")
        return False
    
    # Извлекаем базовое имя файла без расширения
    input_basename = os.path.splitext(os.path.basename(input_file))[0]
    
    # Создаем структуру папок: task/<название_картинки>/ и task/<название_картинки>/a4/
    task_dir = os.path.join("task", input_basename)
    task_a4_dir = os.path.join(task_dir, "A4-pdf")
    
    # Очищаем папку task/<имя файла> перед созданием новых файлов
    if os.path.exists(task_dir):
        print(f"[CLEAN] Очистка папки {task_dir}...")
        try:
            # Удаляем все файлы и подпапки
            for root, dirs, files in os.walk(task_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        print(f"[WARNING] Не удалось удалить {file_path}: {e}")
                for dir_name in dirs:
                    dir_path = os.path.join(root, dir_name)
                    try:
                        os.rmdir(dir_path)
                    except:
                        pass  # Игнорируем ошибки удаления папок (могут быть непустыми)
        except Exception as e:
            print(f"[WARNING] Ошибка при очистке папки {task_dir}: {e}")
    
    os.makedirs(task_dir, exist_ok=True)
    os.makedirs(task_a4_dir, exist_ok=True)
    
    # Пути для сохранения файлов
    output_file = os.path.join(task_dir, f"{input_basename}_layout.jpg")
    
    print(f"[START] Начало обработки:")
    print(f"   Входной файл: {input_file}")
    print(f"   Выходная папка: {task_dir}")
    print(f"   Количество цветов: {num_colors}")
    print(f"   Количество блоков: {num_blocks_width}x{num_blocks_height}")
    print()
    
    try:
        # Создаем раскладку
        palette = create_color_layout(input_file, output_file, num_colors, num_blocks_width, num_blocks_height)
        
        # Нормализуем палитру до кортежей для единообразия
        normalized_palette = []
        for color in palette:
            if isinstance(color, (list, np.ndarray)):
                normalized_color = tuple(int(c) for c in color)
            elif isinstance(color, tuple):
                normalized_color = tuple(int(c) for c in color)
            else:
                normalized_color = color
            normalized_palette.append(normalized_color)
        palette = normalized_palette
        
        # Создаем соответствие между цветами и символами для палитры и органайзера
        char_list = SPECIAL_CHARS * ((len(palette) // len(SPECIAL_CHARS)) + 1)
        color_to_char = {color: char_list[i] for i, color in enumerate(palette)}
        
        # Рассчитываем количество кубиков для каждого цвета (для столбца Длина)
        color_counts = calculate_color_counts(output_file, palette, num_blocks_width, num_blocks_height)
        
        # Сохраняем данные в Excel файл <артикул>.xlsx
        from utils.path_utils import get_static_path
        gamma_excel_path = get_static_path("DMCtoGamma_with_Gamma_OFF_formattedColor.xlsx")
        gamma_legend_path = os.path.join(task_dir, f"{input_basename}.xlsx")
        save_gamma_legend_to_excel(palette, color_to_char, gamma_excel_path if os.path.exists(gamma_excel_path) else None, gamma_legend_path, color_counts)
        
        # Создаем PDF таблицу (передаем ту же палитру, что использовалась для layout)
        pdf_output = os.path.join(task_dir, f"{input_basename}_table.pdf")
        create_pdf_table(output_file, pdf_output, num_blocks_width, num_blocks_height, include_symbols=False, palette=palette)
        
        # Создаем PDF таблицу с символами (передаем ту же палитру)
        pdf_symbols_output = os.path.join(task_dir, f"{input_basename}_table_symbols.pdf")
        create_pdf_table(output_file, pdf_symbols_output, num_blocks_width, num_blocks_height, include_symbols=True, palette=palette)
        
        # Создаем PDF таблицу только с символами (без цветных блоков) (передаем ту же палитру)
        pdf_symbols_only_output = os.path.join(task_dir, f"{input_basename}_table_symbols_only.pdf")
        create_pdf_symbols_only(output_file, pdf_symbols_only_output, num_blocks_width, num_blocks_height, palette=palette)
        
        # Загружаем color_to_char из Excel для создания палитры в правильном порядке
        gamma_legend_path = os.path.join(task_dir, f"{input_basename}.xlsx")
        palette_numbers = None  # Номера N из Excel
        if os.path.exists(gamma_legend_path):
            loaded_color_to_char, loaded_palette, loaded_numbers, loaded_dmc_numbers = load_gamma_legend_from_excel(gamma_legend_path)
            if loaded_color_to_char and loaded_palette:
                color_to_char = loaded_color_to_char
                # Используем палитру из Excel, отсортированную по нумерации N
                palette = loaded_palette
                palette_numbers = loaded_numbers
        
        # Создаем PDF палитры с символами (в порядке нумерации из Excel)
        palette_pdf_path = os.path.join(task_dir, f"{input_basename}_layout_palette.pdf")
        create_palette_pdf(palette, color_to_char, palette_pdf_path, numbers=palette_numbers)
        
        # Конвертируем PDF в JPG
        palette_image_path = os.path.join(task_dir, f"{input_basename}_layout_palette.jpg")
        if not pdf_to_jpg(palette_pdf_path, palette_image_path, dpi=300):
            # Если конвертация не удалась, создаем JPG напрямую (fallback)
            print(f"[WARNING] Не удалось конвертировать PDF в JPG, создаем JPG напрямую...")
            create_palette_image(palette, color_to_char, palette_image_path, numbers=palette_numbers)
        
        # Добавляем страницу с таблицей из Excel в PDF файлы (table_symbols.pdf и table_symbols_only.pdf)
        print(f"[TABLE] Добавление таблицы из Excel в PDF файлы...")
        if os.path.exists(gamma_legend_path):
            add_table_page_to_pdf(pdf_symbols_output, gamma_legend_path, colorize_symbols=True)
            add_table_page_to_pdf(pdf_symbols_only_output, gamma_legend_path, colorize_symbols=False)
        else:
            print(f"[WARNING] Excel файл не найден: {gamma_legend_path}, пропускаем добавление таблицы")
        
        # Создаем органайзер с сопоставлением Gamma цветов
        organizer_path = os.path.join(task_dir, f"{input_basename}_Органайзер.jpg")
        from utils.path_utils import get_static_path
        gamma_excel_path = get_static_path("DMCtoGamma_with_Gamma_OFF_formattedColor.xlsx")
        create_organizer_image(palette, color_to_char, organizer_path, project_name=input_basename, gamma_excel_path=gamma_excel_path if os.path.exists(gamma_excel_path) else None, layout_image_path=output_file, num_blocks_width=num_blocks_width, num_blocks_height=num_blocks_height)
        
        # Очищаем папку a4 перед созданием новых файлов
        if os.path.exists(task_a4_dir):
            for filename in os.listdir(task_a4_dir):
                file_path = os.path.join(task_a4_dir, filename)
                try:
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                except Exception as e:
                    print(f"[WARNING] Не удалось удалить {file_path}: {e}")
        
        # Создаем страницы A4 для PDF с символами и цветами (передаем ту же палитру и путь к палитре)
        base_name_symbols = f"{input_basename}_table_symbols"
        total_pages_colored = create_a4_pages(output_file, base_name_symbols, num_blocks_width, num_blocks_height, 
                        include_colors=True, include_symbols=True, blocks_per_page_width=None, blocks_per_page_height=None, output_dir=task_a4_dir, palette=palette, palette_image_path=palette_image_path, project_name=input_basename)
        
        # Создаем страницы A4 для PDF только с символами (передаем ту же палитру и путь к палитре)
        base_name_symbols_only = f"{input_basename}_table_symbols_only"
        total_pages_bw = create_a4_pages(output_file, base_name_symbols_only, num_blocks_width, num_blocks_height, 
                        include_colors=False, include_symbols=True, blocks_per_page_width=None, blocks_per_page_height=None, output_dir=task_a4_dir, palette=palette, palette_image_path=palette_image_path, project_name=input_basename)
        
        # Объединяем цветные страницы A4 в один PDF
        base_filename_symbols = base_name_symbols
        merged_colored_path = os.path.join(task_a4_dir, f"{base_filename_symbols}_merged.pdf")
        merge_pdf_pages(task_a4_dir, base_filename_symbols, total_pages_colored, merged_colored_path)
        
        # Объединяем черно-белые страницы A4 в один PDF
        base_filename_symbols_only = base_name_symbols_only
        merged_bw_path = os.path.join(task_a4_dir, f"{base_filename_symbols_only}_merged.pdf")
        merge_pdf_pages(task_a4_dir, base_filename_symbols_only, total_pages_bw, merged_bw_path)
        
        # Создание A5 главной страницы (используем исходное изображение проекта)
        print()
        print(f"[INFO] Создание главной страницы...")
        
        # Ищем изображение для главной страницы (приоритет: исходное изображение проекта)
        project_image_path = None
        
        # Вариант 1: Исходное изображение проекта (то, с которым работаем)
        if os.path.exists(input_file):
            project_image_path = input_file
            print(f"[INFO] Используется исходное изображение проекта: {project_image_path}")
        
        # Вариант 2: Layout изображение проекта (если исходное не найдено)
        if not project_image_path:
            layout_image = os.path.join(task_dir, f"{input_basename}_layout.jpg")
            if os.path.exists(layout_image):
                project_image_path = layout_image
                print(f"[INFO] Используется layout изображение: {project_image_path}")
        
        # Вариант 3: Общее A5 изображение из sheet/ (fallback)
        if not project_image_path:
            sheet_dir = "sheet"
            if os.path.exists(sheet_dir):
                # Ищем общий файл a5.jpg
                common_a5 = os.path.join(sheet_dir, "a5.jpg")
                if os.path.exists(common_a5):
                    project_image_path = common_a5
                    print(f"[INFO] Используется общее A5 изображение: {project_image_path}")
                else:
                    # Ищем файл с именем проекта
                    project_a5 = os.path.join(sheet_dir, f"{input_basename}_a5.png")
                    if os.path.exists(project_a5):
                        project_image_path = project_a5
                        print(f"[INFO] Используется A5 изображение проекта: {project_image_path}")
                    else:
                        # Ищем любой файл с a5 в имени
                        for f in os.listdir(sheet_dir):
                            if 'a5' in f.lower() and f.lower().endswith(('.png', '.jpg', '.jpeg')):
                                project_image_path = os.path.join(sheet_dir, f)
                                print(f"[INFO] Используется A5 изображение: {project_image_path}")
                                break
        
        if project_image_path and os.path.exists(project_image_path):
            try:
                from create_a5_main_from_image import create_a5_main_page_from_image
                
                # Извлекаем название файла изображения (без расширения) для использования в баре
                image_filename = os.path.splitext(os.path.basename(project_image_path))[0]
                # Убираем суффиксы _layout, _a5 если есть
                if image_filename.endswith('_layout'):
                    image_filename = image_filename[:-7]
                elif image_filename.endswith('_a5') or image_filename.endswith('_A5'):
                    image_filename = image_filename[:-3]
                
                # Используем название файла как артикул для бара
                article_text = image_filename
                print(f"[INFO] Артикул для бара (из названия файла): {article_text}")
                
                # Создаем главную страницу в папке task/<название>/
                main_page_path = create_a5_main_page_from_image(
                    image_path=project_image_path,
                    pdf_name=input_basename,
                    article_text=article_text,
                    output_folder=task_dir,  # Сохраняем в папку проекта
                    num_colors=len(palette)  # Передаем количество цветов из палитры
                )
                print(f"[SUCCESS] Создана A5 главная страница: {main_page_path}")
            except Exception as e:
                print(f"[WARNING] Ошибка при создании главной страницы: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"[WARNING] Изображение для главной страницы не найдено, пропускаем создание")
            print(f"[INFO] Искали:")
            print(f"  - Исходное изображение: {input_file}")
            print(f"  - Layout изображение: {task_dir}/{input_basename}_layout.jpg")
            print(f"  - Общее A5: sheet/a5.jpg")
        
        print()
        print(f"[SUCCESS] Обработка завершена успешно!")
        print(f"   Раскладка: {output_file}")
        print(f"   Палитра: {palette_image_path}")
        print(f"   Органайзер: {organizer_path}")
        print(f"   PDF таблица: {pdf_output}")
        print(f"   PDF таблица с символами (с палитрой): {pdf_symbols_output}")
        print(f"   PDF таблица только с символами (с палитрой): {pdf_symbols_only_output}")
        print(f"   Объединенный PDF (цветной): {merged_colored_path}")
        print(f"   Объединенный PDF (черно-белый): {merged_bw_path}")
        print(f"   Отдельные страницы A4: {task_a4_dir}/{base_filename_symbols}_a4_page_*.pdf")
        print(f"   Отдельные страницы A4 (ЧБ): {task_a4_dir}/{base_filename_symbols_only}_a4_page_*.pdf")
        print(f"   Использовано цветов: {len(palette)}")
        print(f"   Количество блоков: {num_blocks_width}x{num_blocks_height}")
        
        # Создаем файл data.json с общей информацией о проекте
        print()
        print(f"[INFO] Создание файла {input_basename}_data.json...")
        try:
            data_file_path = os.path.join(task_dir, f"{input_basename}_data.json")
            # Получаем размер вышивки из config
            embroidery_size = "Не указан"
            if hasattr(config, 'EMBROIDERY_SETTINGS') and config.EMBROIDERY_SETTINGS.get('size'):
                embroidery_size = config.EMBROIDERY_SETTINGS.get('size', 'Не указан')
            
            # Формируем данные в виде словаря
            data = {
                "название_проекта": input_basename,
                "артикул": input_basename,
                "количество_цветов": len(palette),
                "размер_вышивки": embroidery_size,
                "количество_ячеек": {
                    "ширина": num_blocks_width,
                    "высота": num_blocks_height
                },
                "тип_цвета": "Gamma"  # По умолчанию Gamma, так как process_single_image не использует диалог конфигурации
            }
            
            # Записываем данные в JSON файл
            with open(data_file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
            print(f"[SUCCESS] Файл {input_basename}_data.json создан: {data_file_path}")
        except Exception as e:
            print(f"[WARNING] Ошибка при создании {input_basename}_data.json: {e}")
            import traceback
            traceback.print_exc()
        
        return True
        
    except Exception as e:
        print(f"[ERROR] Ошибка при обработке: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Основная функция"""
    # Параметры по умолчанию из конфига
    num_colors = config.NUM_COLORS
    num_blocks_width = config.NUM_BLOCKS_WIDTH
    num_blocks_height = config.NUM_BLOCKS_HEIGHT
    
    # Проверяем аргументы командной строки
    if len(sys.argv) > 1:
        # Если указан файл - обрабатываем только его
        input_file = sys.argv[1]
        if len(sys.argv) > 2:
            num_colors = int(sys.argv[2])
        if len(sys.argv) > 3:
            num_blocks_width = int(sys.argv[3])
        if len(sys.argv) > 4:
            num_blocks_height = int(sys.argv[4])
        
        # Обрабатываем указанный файл
        process_single_image(input_file, num_colors, num_blocks_width, num_blocks_height)
    else:
        # Если аргументов нет - обрабатываем все изображения из папки sheet/
        sheet_dir = config.SHEET_DIR
        
        if not os.path.exists(sheet_dir):
            print(f"[ERROR] Папка {sheet_dir} не найдена!")
            print(f"[INFO] Создайте папку {sheet_dir}/ и поместите туда изображения для обработки")
            return
        
        # Ищем все изображения в папке sheet/
        image_extensions = ('.png', '.jpg', '.jpeg', '.PNG', '.JPG', '.JPEG', '.webp', '.WEBP')
        image_files = []
        
        for filename in os.listdir(sheet_dir):
            file_path = os.path.join(sheet_dir, filename)
            if os.path.isfile(file_path) and filename.endswith(image_extensions):
                # Пропускаем файлы с a5 в имени (это специальные файлы для главной страницы)
                if 'a5' not in filename.lower():
                    image_files.append(file_path)
        
        if not image_files:
            print(f"[INFO] Изображения для обработки не найдены в папке {sheet_dir}/")
            print(f"[INFO] Поместите изображения (.png, .jpg, .jpeg, .webp) в папку {sheet_dir}/")
            return
        
        print(f"[INFO] Найдено изображений для обработки: {len(image_files)}")
        print(f"[INFO] Изображения: {[os.path.basename(f) for f in image_files]}")
        print()
        
        # Обрабатываем каждое изображение
        successful = 0
        failed = 0
        
        for i, image_file in enumerate(image_files, 1):
            print(f"\n{'=' * 80}")
            print(f"Обработка {i}/{len(image_files)}: {os.path.basename(image_file)}")
            print(f"{'=' * 80}")
            
            if process_single_image(image_file, num_colors, num_blocks_width, num_blocks_height):
                successful += 1
            else:
                failed += 1
        
        print(f"\n{'=' * 80}")
        print(f"Обработка завершена!")
        print(f"  Успешно обработано: {successful}")
        print(f"  Ошибок: {failed}")
        print(f"{'=' * 80}")


if __name__ == "__main__":
    main()

